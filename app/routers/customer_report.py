import asyncio
import json
import os
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from app.db import get_db
from app.models import User, CustomerReport, AuditLog
from app.schemas import CustomerReportCreate, CustomerReportUpdate, CustomerReportDto
from app.deps import require_admin

router = APIRouter(tags=["Customer Report"])


@router.get("/customer-report", response_model=list[CustomerReportDto])
async def get_customer_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Получить все отчеты по клиентам"""
    result = await db.execute(
        select(CustomerReport).order_by(CustomerReport.created_date.desc())
    )
    reports = result.scalars().all()
    return [CustomerReportDto.model_validate(report, from_attributes=True) for report in reports]


@router.get("/customer-report/export/excel", summary="Экспорт отчета по клиентам в Excel")
async def export_customer_report_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Выгрузка всех записей CustomerReport в Excel по шаблону"""
    result = await db.execute(
        select(CustomerReport).order_by(CustomerReport.created_date.desc())
    )
    reports = result.scalars().all()

    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "template", "CustomerReport.xlsx"
    )
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template file not found")

    output = await asyncio.to_thread(_build_customer_report_excel, reports, template_path)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CustomerReport.xlsx"}
    )


def _build_customer_report_excel(reports, template_path) -> BytesIO:
    wb = load_workbook(template_path)
    ws = wb["Data Nasabah"]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for idx, report in enumerate(reports, start=1):
        row = idx + 3  # data starts at row 4 (row 3 is header)

        values = [
            idx,                                                          # A: No.
            report.registration_number,                                   # B: No.Identitas
            report.name,                                                  # C: Nama Pengirim
            report.birth_place_date,                                      # D: Tempat dan Tanggal Lahir
            report.address,                                               # E: Alamat Pengirim
            report.tax_number,                                            # F: No.NPWP
            report.occupation,                                            # G: Pekerjaan
            "WNI" if report.indonesian_citizenship else "WNA",            # H: Kewarganegaraan
            report.gender,                                                # I: Jenis Kelamin
            report.phone_number,                                          # J: No. Telepon
            report.recipient_name,                                        # K: Nama Penerima
            report.recipient_address,                                     # L: Alamat Penerima
            report.legal_tax_number,                                      # M: No.Identitas Lain
            "Ya" if report.pep_indicator else "Tidak",                    # N: Indikator PEP
            report.code_type,                                             # O: Kode MC/Bank
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/customer-report", response_model=CustomerReportDto, status_code=status.HTTP_201_CREATED)
async def create_customer_report(
    data: CustomerReportCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Создать новый отчет по клиенту"""
    report = CustomerReport(
        **data.model_dump(),
        created_date=datetime.utcnow(),
        created_by=current_user.username
    )
    db.add(report)
    
    audit_log = AuditLog(
        entity="customer_reports",
        entity_id=None,
        action="CREATE",
        old_value=None,
        new_value=json.dumps(data.model_dump(), ensure_ascii=False, default=str),
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.commit()
    await db.refresh(report)
    
    audit_log.entity_id = str(report.id)
    await db.commit()
    
    return CustomerReportDto.model_validate(report, from_attributes=True)


@router.put("/customer-report/{report_id}", response_model=CustomerReportDto)
async def update_customer_report(
    report_id: int,
    data: CustomerReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Обновить отчет по клиенту"""
    result = await db.execute(
        select(CustomerReport).where(CustomerReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Customer report with id {report_id} not found"
        )
    
    old_value = {
        "customer_type": report.customer_type,
        "registration_number": report.registration_number,
        "tax_number": report.tax_number,
        "legal_tax_number_type": report.legal_tax_number_type,
        "legal_tax_number": report.legal_tax_number,
        "name": report.name,
        "address": report.address,
        "indonesian_citizenship": report.indonesian_citizenship,
        "director_name": report.director_name,
        "pep_indicator": report.pep_indicator,
        "code_type": report.code_type,
        "business_area": report.business_area
    }
    
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(report, field, value)
    
    report.updated_at = datetime.utcnow()
    report.updated_by = current_user.username
    
    new_value = old_value.copy()
    new_value.update(update_data)
    
    audit_log = AuditLog(
        entity="customer_reports",
        entity_id=str(report_id),
        action="UPDATE",
        old_value=json.dumps(old_value, ensure_ascii=False, default=str),
        new_value=json.dumps(new_value, ensure_ascii=False, default=str),
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.commit()
    await db.refresh(report)
    
    return CustomerReportDto.model_validate(report, from_attributes=True)


@router.delete("/customer-report/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_customer_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Удалить отчет по клиенту"""
    result = await db.execute(
        select(CustomerReport).where(CustomerReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Customer report with id {report_id} not found"
        )
    
    old_value = {
        "id": report.id,
        "customer_type": report.customer_type,
        "name": report.name,
        "registration_number": report.registration_number
    }
    
    audit_log = AuditLog(
        entity="customer_reports",
        entity_id=str(report_id),
        action="DELETE",
        old_value=json.dumps(old_value, ensure_ascii=False, default=str),
        new_value=None,
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.delete(report)
    await db.commit()

    return None
