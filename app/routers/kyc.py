import asyncio
from datetime import datetime
from typing import List, Optional
from io import BytesIO
import os
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import attributes
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from app.db import get_db
from app.deps import get_current_user
from app.models import User, Client, OnboardingKycProfile, OnboardingKycStatusHistory, OnboardingKycUbo, OnboardingKycDocument, CustomerReport, AuditLog
import json as _json
from app.enums import KYCStatus, KYCDocumentType
from app.s3_client import S3Client, get_s3_client, generate_kyc_s3_key
from app.file_validators import validate_pdf_not_encrypted
from app.logger import setup_logger
from app.config import settings
from app.schemas import (
    KYCProfileResponse, 
    KYCProfileUpdateRequest, 
    KYCSubmitResponse,
    KYCProfilePayload,
    KYCCorporateDetailsDto,
    KYCBankingDetailsDto,
    KYCDeclarationDto,
    UBOCreateRequest,
    UBOUpdateRequest,
    UBOResponse,
    KYCDocumentDto,
    KYCDocumentUploadResponse,
    PresignedUrlResponse,
    KYCQueueItemDto,
    KYCDecisionRequest,
    KYCDecisionResponse
)

router = APIRouter(prefix="/api/v1", tags=["KYC"])

# Инициализация логгера для KYC роутера
logger = setup_logger("kyc_router", level=settings.LOG_LEVEL, debug=settings.DEBUG)


async def _check_client_access(db: AsyncSession, client_id: str, current_user: User):
    result = await db.execute(select(Client).where(Client.client_id == client_id))
    client = result.scalar_one_or_none()
    
    if not client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Клиент с ID {client_id} не найден"
        )
    
    if current_user.role == "CLIENT" and client.user_id != current_user.user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Нет доступа к данному клиенту"
        )
    
    return client


def _payload_to_dto(payload: dict) -> KYCProfilePayload:
    corporate = None
    banking = None
    declaration = None
    
    if payload.get("corporate") is not None:
        corporate = KYCCorporateDetailsDto(**payload["corporate"])
    
    if payload.get("banking") is not None:
        banking = KYCBankingDetailsDto(**payload["banking"])
    
    if payload.get("declaration") is not None:
        declaration = KYCDeclarationDto(**payload["declaration"])
    
    return KYCProfilePayload(
        corporate=corporate,
        banking=banking,
        declaration=declaration
    )


@router.get("/clients/{client_id}/kyc", response_model=KYCProfileResponse)
async def get_kyc_profile(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)

    payload_data = profile.payload or {}
    
    return KYCProfileResponse(
        profile_id=profile.profile_id,
        client_id=profile.client_id,
        status=profile.status,
        version=profile.version,
        data=_payload_to_dto(payload_data),
        created_at=profile.created_at,
        updated_at=profile.updated_at,
        submitted_at=profile.submitted_at,
        decided_at=profile.decided_at,
        decided_by=profile.decided_by,
        decision_comment=profile.decision_comment
    )


@router.put("/clients/{client_id}/kyc", response_model=KYCProfileResponse)
async def update_kyc_profile(
    client_id: str,
    data: KYCProfileUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    client = await _check_client_access(db, client_id, current_user)
    
    result = await db.execute(
        select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
    )
    profile = result.scalar_one_or_none()
    
    now = datetime.utcnow()
    
    if not profile:
        profile = OnboardingKycProfile(
            client_id=client_id,
            status=KYCStatus.IN_PROGRESS.value,
            payload={},
            version=1,
            created_at=now,
            updated_at=now
        )
        db.add(profile)
        await db.flush()
    
    if profile.status not in [KYCStatus.CREATED.value, KYCStatus.IN_PROGRESS.value, KYCStatus.NEEDS_FIX.value, KYCStatus.APPROVED.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невозможно редактировать KYC в статусе '{profile.status}'"
        )
    
    payload = profile.payload or {}
    
    corporate = payload.get("corporate", {})
    if data.company_name is not None:
        corporate["company_name"] = data.company_name
    if data.trading_name is not None:
        corporate["trading_name"] = data.trading_name
    if data.incorporation_date is not None:
        corporate["incorporation_date"] = data.incorporation_date.isoformat()
    if data.incorporation_country is not None:
        corporate["incorporation_country"] = data.incorporation_country
    if data.registration_number is not None:
        corporate["registration_number"] = data.registration_number
    if data.tax_id is not None:
        corporate["tax_id"] = data.tax_id
    if data.registered_address is not None:
        corporate["registered_address"] = data.registered_address
    if data.telephone is not None:
        corporate["telephone"] = data.telephone
    if data.website is not None:
        corporate["website"] = data.website
    
    banking = payload.get("banking", {})
    if data.principal_bankers is not None:
        banking["principal_bankers"] = data.principal_bankers
    if data.swift_bic is not None:
        banking["swift_bic"] = data.swift_bic
    if data.bank_branch_address is not None:
        banking["bank_branch_address"] = data.bank_branch_address
    if data.bank_city_country is not None:
        banking["bank_city_country"] = data.bank_city_country
    if data.bank_account_name is not None:
        banking["bank_account_name"] = data.bank_account_name
    if data.bank_account_currency is not None:
        banking["bank_account_currency"] = data.bank_account_currency
    if data.bank_account_number is not None:
        banking["bank_account_number"] = data.bank_account_number
    if data.bank_manager_contact is not None:
        banking["bank_manager_contact"] = data.bank_manager_contact
    
    declaration = payload.get("declaration", {})
    if data.declaration_confirmed is not None:
        declaration["declaration_confirmed"] = data.declaration_confirmed
    if data.authorized_person_name is not None:
        declaration["authorized_person_name"] = data.authorized_person_name
    if data.signature_date is not None:
        declaration["signature_date"] = data.signature_date.isoformat()
    if data.authorized_person_position is not None:
        declaration["authorized_person_position"] = data.authorized_person_position
    if data.signature_location is not None:
        declaration["signature_location"] = data.signature_location
    if data.signed_kyc_document_url is not None:
        declaration["signed_kyc_document_url"] = data.signed_kyc_document_url
    
    payload["corporate"] = corporate
    payload["banking"] = banking
    payload["declaration"] = declaration
    
    old_status = profile.status
    profile.payload = payload
    
    # КРИТИЧНО: SQLAlchemy не отслеживает изменения внутри JSON полей
    # Нужно явно указать что поле изменилось
    attributes.flag_modified(profile, "payload")
    
    profile.status = KYCStatus.IN_PROGRESS.value
    profile.updated_at = now
    
    client.kyc_status = KYCStatus.IN_PROGRESS.value
    
    if old_status != KYCStatus.IN_PROGRESS.value:
        history = OnboardingKycStatusHistory(
            profile_id=profile.profile_id,
            old_status=old_status,
            new_status=KYCStatus.IN_PROGRESS.value,
            changed_by=current_user.user_id,
            changed_at=now,
            comment="KYC профиль обновлен"
        )
        db.add(history)
    
    await db.commit()
    await db.refresh(profile)
    
    return KYCProfileResponse(
        profile_id=profile.profile_id,
        client_id=profile.client_id,
        status=profile.status,
        version=profile.version,
        data=_payload_to_dto(profile.payload),
        created_at=profile.created_at,
        updated_at=profile.updated_at,
        submitted_at=profile.submitted_at,
        decided_at=profile.decided_at,
        decided_by=profile.decided_by,
        decision_comment=profile.decision_comment
    )


@router.post("/clients/{client_id}/kyc/submit", response_model=KYCSubmitResponse)
async def submit_kyc(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    client = await _check_client_access(db, client_id, current_user)
    
    result = await db.execute(
        select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
    )
    profile = result.scalar_one_or_none()
    
    if not profile:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"KYC профиль для клиента {client_id} не найден. Сначала заполните данные."
        )
    
    if profile.status not in [KYCStatus.IN_PROGRESS.value, KYCStatus.NEEDS_FIX.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невозможно отправить KYC в статусе '{profile.status}'"
        )
    
    now = datetime.utcnow()
    old_status = profile.status
    
    profile.status = KYCStatus.SUBMITTED.value
    profile.submitted_at = now
    profile.updated_at = now
    
    client.kyc_status = KYCStatus.SUBMITTED.value
    client.kyc_submitted_at = now
    
    history = OnboardingKycStatusHistory(
        profile_id=profile.profile_id,
        old_status=old_status,
        new_status=KYCStatus.SUBMITTED.value,
        changed_by=current_user.user_id,
        changed_at=now,
        comment="KYC отправлен на проверку"
    )
    db.add(history)

    # ТЗ Sec 4.12: дублируем в глобальный audit_log, чтобы регулятор
    # мог найти все sensitive-события через одну таблицу. История в
    # OnboardingKycStatusHistory остаётся (domain-specific трейс),
    # AuditLog даёт cross-entity compliance-view.
    db.add(AuditLog(
        entity="clients",
        entity_id=client.client_id,
        action="KYC_SUBMITTED",
        old_value=_json.dumps({
            "kyc_status": old_status,
            "profile_id": profile.profile_id,
        }, ensure_ascii=False),
        new_value=_json.dumps({
            "kyc_status": KYCStatus.SUBMITTED.value,
            "profile_id": profile.profile_id,
            "submitted_at": now.isoformat(),
        }, ensure_ascii=False),
        created_by=current_user.username,
        created_at=now,
    ))

    await db.commit()
    await db.refresh(profile)

    return KYCSubmitResponse(
        profile_id=profile.profile_id,
        client_id=profile.client_id,
        status=profile.status,
        submitted_at=profile.submitted_at,
        message="KYC профиль успешно отправлен на проверку"
    )


async def _get_or_create_profile(db: AsyncSession, client_id: str) -> OnboardingKycProfile:
    """Получить или создать KYC профиль для клиента"""
    result = await db.execute(
        select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
    )
    profile = result.scalar_one_or_none()

    if not profile:
        try:
            now = datetime.utcnow()
            profile = OnboardingKycProfile(
                client_id=client_id,
                status=KYCStatus.CREATED.value,
                payload={},
                version=1,
                created_at=now,
                updated_at=now
            )
            db.add(profile)
            await db.commit()
            await db.refresh(profile)
        except IntegrityError:
            await db.rollback()
            result = await db.execute(
                select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
            )
            profile = result.scalar_one_or_none()

    return profile


# ======================================================================
# UBO ENDPOINTS
# ======================================================================

@router.get("/clients/{client_id}/ubos", response_model=List[UBOResponse])
async def list_ubos(
    client_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Получить список всех UBO для клиента"""
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    result = await db.execute(
        select(OnboardingKycUbo).where(OnboardingKycUbo.profile_id == profile.profile_id)
    )
    ubos = result.scalars().all()
    
    return ubos


@router.post("/clients/{client_id}/ubos", response_model=UBOResponse, status_code=status.HTTP_201_CREATED)
async def create_ubo(
    client_id: str,
    data: UBOCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Создать нового UBO для клиента"""
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    now = datetime.utcnow()
    
    ubo = OnboardingKycUbo(
        profile_id=profile.profile_id,
        ubo_name=data.ubo_name,
        shareholding_percent=data.shareholding_percent,
        nationality=data.nationality,
        residence_country=data.residence_country,
        created_at=now,
        updated_at=now
    )
    
    db.add(ubo)
    await db.commit()
    await db.refresh(ubo)
    
    return ubo


@router.put("/clients/{client_id}/ubos/{ubo_id}", response_model=UBOResponse)
async def update_ubo(
    client_id: str,
    ubo_id: int,
    data: UBOUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Обновить данные UBO"""
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    result = await db.execute(
        select(OnboardingKycUbo).where(
            OnboardingKycUbo.id == ubo_id,
            OnboardingKycUbo.profile_id == profile.profile_id
        )
    )
    ubo = result.scalar_one_or_none()
    
    if not ubo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"UBO с ID {ubo_id} не найден для клиента {client_id}"
        )
    
    if data.ubo_name is not None:
        ubo.ubo_name = data.ubo_name
    if data.shareholding_percent is not None:
        ubo.shareholding_percent = data.shareholding_percent
    if data.nationality is not None:
        ubo.nationality = data.nationality
    if data.residence_country is not None:
        ubo.residence_country = data.residence_country
    
    ubo.updated_at = datetime.utcnow()
    
    await db.commit()
    await db.refresh(ubo)
    
    return ubo


@router.delete("/clients/{client_id}/ubos/{ubo_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_ubo(
    client_id: str,
    ubo_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Удалить UBO"""
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    result = await db.execute(
        select(OnboardingKycUbo).where(
            OnboardingKycUbo.id == ubo_id,
            OnboardingKycUbo.profile_id == profile.profile_id
        )
    )
    ubo = result.scalar_one_or_none()
    
    if not ubo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"UBO с ID {ubo_id} не найден для клиента {client_id}"
        )
    
    await db.delete(ubo)
    await db.commit()
    
    return None


# ======================================================================
# KYC DOCUMENTS ENDPOINTS
# ======================================================================

@router.post("/clients/{client_id}/documents", response_model=KYCDocumentUploadResponse)
async def upload_kyc_document(
    client_id: str,
    file: UploadFile = File(...),
    doc_type: str = Form(...),
    comment: Optional[str] = Form(None),
    is_required: bool = Form(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    s3: S3Client = Depends(get_s3_client)
):
    """
    Загрузить KYC документ для клиента
    
    Параметры:
    - file: Файл для загрузки
    - doc_type: Тип документа (например, certificate_of_incorporation, ubo_passport)
    - comment: Комментарий к документу
    - is_required: Является ли документ обязательным
    """
    logger.debug(
        f"KYC document upload request: client_id={client_id}, doc_type={doc_type}, "
        f"filename={file.filename if file else 'None'}, user={current_user.username if current_user else 'None'}"
    )
    
    try:
        await _check_client_access(db, client_id, current_user)
    except HTTPException as e:
        logger.error(f"Client access error: {e.detail}")
        raise
    
    try:
        doc_type_enum = KYCDocumentType(doc_type)
        logger.debug(f"Document type validated: {doc_type_enum}")
    except ValueError as e:
        allowed_types = ', '.join([t.value for t in KYCDocumentType])
        logger.error(f"Invalid document type: {doc_type}. Allowed: {allowed_types}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недопустимый тип документа: {doc_type}. Разрешенные типы: {', '.join([t.value for t in KYCDocumentType])}"
        )
    
    logger.debug(f"Getting or creating KYC profile for client {client_id}")
    profile = await _get_or_create_profile(db, client_id)
    logger.debug(f"KYC profile obtained: profile_id={profile.profile_id}")
    
    if not file.filename:
        logger.error("File name not specified")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Имя файла не указано"
        )
    
    logger.debug(f"File accepted: {file.filename}")
    
    content_type = file.content_type or 'application/octet-stream'
    
    file_content = await file.read()
    file_size = len(file_content)

    max_size = 50 * 1024 * 1024
    if file_size > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Размер файла превышает максимально допустимый ({max_size // 1024 // 1024} МБ)"
        )

    # ТЗ Sec 5.5.3: блокируем PDF с парольной защитой —
    # staff не сможет их открыть для проверки KYC.
    await validate_pdf_not_encrypted(file_content, file.filename)
    
    # Multi-file per doc_type: всегда создаём новую запись.
    # Лимит — 10 файлов на (profile_id, doc_type). Проверяем ДО загрузки в S3,
    # чтобы не оставлять висячие файлы в storage при отказе.
    count_result = await db.execute(
        select(func.count()).select_from(OnboardingKycDocument).where(
            OnboardingKycDocument.profile_id == profile.profile_id,
            OnboardingKycDocument.doc_type == doc_type,
        )
    )
    existing_count = count_result.scalar_one()
    if existing_count >= 10:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Лимит достигнут: максимум 10 файлов для типа {doc_type}",
        )

    s3_key = generate_kyc_s3_key(client_id, doc_type, file.filename)

    await s3.upload_file(
        file=BytesIO(file_content),
        key=s3_key,
        content_type=content_type
    )

    now = datetime.utcnow()
    new_doc = OnboardingKycDocument(
        profile_id=profile.profile_id,
        doc_type=doc_type,
        file_key=s3_key,
        file_url=s3_key,
        file_name=file.filename,
        file_size=file_size,
        uploaded_by=current_user.user_id,
        uploaded_at=now,
        is_required=is_required,
        comment=comment,
    )
    db.add(new_doc)
    await db.flush()
    doc_id = new_doc.doc_id

    await db.commit()
    
    return KYCDocumentUploadResponse(
        doc_id=doc_id,
        doc_type=doc_type,
        file_name=file.filename,
        file_size=file_size,
        uploaded_at=now,
        message="KYC документ успешно загружен"
    )


@router.get("/clients/{client_id}/documents", response_model=List[KYCDocumentDto])
async def list_kyc_documents(
    client_id: str,
    doc_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Получить список KYC документов для клиента
    
    Параметры:
    - doc_type: Фильтр по типу документа (опционально)
    """
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    query = select(OnboardingKycDocument).where(
        OnboardingKycDocument.profile_id == profile.profile_id
    )
    
    if doc_type:
        try:
            KYCDocumentType(doc_type)
            query = query.where(OnboardingKycDocument.doc_type == doc_type)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Недопустимый тип документа: {doc_type}"
            )
    
    query = query.order_by(OnboardingKycDocument.uploaded_at.desc())
    result = await db.execute(query)
    documents = result.scalars().all()
    
    return documents


@router.get("/clients/{client_id}/documents/{doc_id}/download", response_model=PresignedUrlResponse)
async def download_kyc_document(
    client_id: str,
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    s3: S3Client = Depends(get_s3_client)
):
    """
    Получить presigned URL для скачивания KYC документа
    
    Параметры:
    - client_id: ID клиента
    - doc_id: ID документа
    
    Возвращает:
    - presigned_url: Временный URL для скачивания (действителен 15 минут)
    - expires_in: Время жизни URL в секундах
    - file_name: Имя файла
    """
    await _check_client_access(db, client_id, current_user)
    
    profile = await _get_or_create_profile(db, client_id)
    
    result = await db.execute(
        select(OnboardingKycDocument).where(
            OnboardingKycDocument.doc_id == doc_id,
            OnboardingKycDocument.profile_id == profile.profile_id
        )
    )
    document = result.scalar_one_or_none()
    
    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"KYC документ с ID {doc_id} не найден для клиента {client_id}"
        )
    
    expiration = 900
    presigned_url = await s3.generate_presigned_url(
        key=document.file_key,
        expiration=expiration
    )
    
    return PresignedUrlResponse(
        presigned_url=presigned_url,
        expires_in=expiration,
        file_name=document.file_name,
        message=f"Presigned URL сгенерирован. Действителен 15 минут."
    )


@router.delete(
    "/clients/{client_id}/documents/{doc_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_kyc_document(
    client_id: str,
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    s3: S3Client = Depends(get_s3_client),
):
    """
    Удалить KYC документ.

    Запрещено после approved/in_progress (профиль уже на проверке или одобрен).
    Разрешено в статусах draft / submitted / needs_fix / rejected, чтобы клиент
    мог исправить ошибочно загруженные файлы перед / после доработки.
    """
    await _check_client_access(db, client_id, current_user)
    profile = await _get_or_create_profile(db, client_id)

    result = await db.execute(
        select(OnboardingKycDocument).where(
            OnboardingKycDocument.doc_id == doc_id,
            OnboardingKycDocument.profile_id == profile.profile_id,
        )
    )
    document = result.scalar_one_or_none()
    if not document:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"KYC документ {doc_id} не найден",
        )

    # Запрет только для approved (финальный иммутабельный статус).
    # in_progress — это "клиент заполняет KYC", тут как раз и нужно делать
    # правки. submitted/needs_fix/rejected — клиент может переделать.
    if profile.status == "approved":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Нельзя удалять документы после одобрения",
        )

    # S3-delete — best effort: даже если файл уже отсутствует, продолжаем
    # удалять запись из БД, чтобы не оставлять висячую строку.
    try:
        if document.file_key:
            await s3.delete_file(document.file_key)
    except Exception as e:
        logger.warning(f"S3 delete failed for {document.file_key}: {e}")

    await db.delete(document)
    await db.commit()


# ======================================================================
# ADMIN KYC ENDPOINTS
# ======================================================================

@router.get("/kyc/queue", response_model=List[KYCQueueItemDto])
async def get_kyc_queue(
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Получить очередь (список) заявок KYC для проверки (только для админов/staff)
    
    Параметры:
    - status: Фильтр по статусу (submitted, in_progress, approved, rejected, needs_fix). 
              По умолчанию возвращает submitted и in_progress.
    
    Возвращает:
    - Список заявок с базовой информацией о клиенте и компании
    """
    if current_user.role not in ["ADMIN", "STAFF", "KYC_OPERATOR"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Доступ запрещен. Требуется роль ADMIN, STAFF или KYC_OPERATOR"
        )
    
    query = select(OnboardingKycProfile, Client, User).join(
        Client, OnboardingKycProfile.client_id == Client.client_id
    ).join(
        User, Client.user_id == User.user_id
    )
    
    if status:
        valid_statuses = [s.value for s in KYCStatus]
        if status not in valid_statuses:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Недопустимый статус: {status}. Разрешенные: {', '.join(valid_statuses)}"
            )
        query = query.where(OnboardingKycProfile.status == status)
    else:
        query = query.where(OnboardingKycProfile.status.in_([
            KYCStatus.SUBMITTED.value, 
            KYCStatus.IN_PROGRESS.value
        ]))
    
    query = query.order_by(OnboardingKycProfile.submitted_at.asc())
    
    result = await db.execute(query)
    rows = result.all()
    
    queue_items = []
    for profile, client, user in rows:
        company_name = None
        if profile.payload and profile.payload.get("corporate"):
            company_name = profile.payload["corporate"].get("company_name")
        
        queue_items.append(KYCQueueItemDto(
            profile_id=profile.profile_id,
            client_id=profile.client_id,
            company_name=company_name,
            client_name=client.client_name,
            client_email=user.email or client.client_mail,
            submitted_at=profile.submitted_at,
            status=profile.status
        ))
    
    return queue_items


@router.post("/clients/{client_id}/kyc/decision", response_model=KYCDecisionResponse)
async def make_kyc_decision(
    client_id: str,
    decision: KYCDecisionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Принять решение по KYC заявке (approve/reject) - только для админов/staff
    
    Параметры:
    - client_id: ID клиента
    - decision: Решение (status: "approved" или "rejected", comment: комментарий)
    
    Возвращает:
    - Информацию о принятом решении
    """
    if current_user.role not in ["ADMIN", "STAFF", "KYC_OPERATOR"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Доступ запрещен. Требуется роль ADMIN, STAFF или KYC_OPERATOR"
        )
    
    if decision.status not in [KYCStatus.APPROVED.value, KYCStatus.REJECTED.value, KYCStatus.NEEDS_FIX.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недопустимый статус решения: {decision.status}. Разрешены: approved, rejected, needs_fix"
        )
    
    if decision.status in [KYCStatus.REJECTED.value, KYCStatus.NEEDS_FIX.value] and not decision.comment:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Комментарий обязателен при статусе '{decision.status}'"
        )
    
    result = await db.execute(
        select(Client).where(Client.client_id == client_id)
    )
    client = result.scalar_one_or_none()
    
    if not client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Клиент с ID {client_id} не найден"
        )
    
    result = await db.execute(
        select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
    )
    profile = result.scalar_one_or_none()
    
    if not profile:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"KYC профиль для клиента {client_id} не найден"
        )
    
    if profile.status not in [KYCStatus.SUBMITTED.value, KYCStatus.IN_PROGRESS.value, KYCStatus.NEEDS_FIX.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невозможно принять решение по KYC в статусе '{profile.status}'. Должен быть submitted, in_progress или needs_fix"
        )
    
    now = datetime.utcnow()
    old_status = profile.status
    
    profile.status = decision.status
    profile.decided_at = now
    profile.decided_by = current_user.user_id
    profile.decision_comment = decision.comment
    profile.updated_at = now
    
    client.kyc_status = decision.status
    client.kyc_decided_at = now
    client.kyc_decided_by = current_user.user_id
    
    history = OnboardingKycStatusHistory(
        profile_id=profile.profile_id,
        old_status=old_status,
        new_status=decision.status,
        changed_by=current_user.user_id,
        changed_at=now,
        comment=decision.comment or f"Решение: {decision.status}"
    )
    db.add(history)

    # ТЗ Sec 4.12: KYC-решения должны попадать в глобальный AuditLog,
    # а не только в OnboardingKycStatusHistory. Action включает суффикс
    # с конкретным решением (APPROVED/REJECTED/NEEDS_FIX), чтобы
    # compliance-офицер мог фильтровать узко через
    # WHERE action='KYC_DECISION_REJECTED' или широко через
    # WHERE action LIKE 'KYC_DECISION_%'.
    db.add(AuditLog(
        entity="clients",
        entity_id=client.client_id,
        action=f"KYC_DECISION_{decision.status.upper()}",
        old_value=_json.dumps({
            "kyc_status": old_status,
            "profile_id": profile.profile_id,
        }, ensure_ascii=False),
        new_value=_json.dumps({
            "kyc_status": decision.status,
            "profile_id": profile.profile_id,
            "decided_at": now.isoformat(),
            "decided_by": current_user.user_id,
            "comment": decision.comment,
        }, ensure_ascii=False),
        created_by=current_user.username,
        created_at=now,
    ))

    if decision.status == KYCStatus.APPROVED.value:
        payload = profile.payload or {}
        corporate = payload.get("corporate", {})

        customer_report = CustomerReport(
            customer_type="client",
            registration_number=client.client_reg_number,
            tax_number=client.client_tax_number,
            legal_tax_number_type="NPWP" if client.client_tax_number else None,
            legal_tax_number=corporate.get("tax_id"),
            name=client.client_name or "Unknown",
            director_name=client.client_director,
            address=corporate.get("registered_address"),
            phone_number=corporate.get("telephone"),
            indonesian_citizenship=(corporate.get("incorporation_country") == "ID"),
            code_type="Corporate",
            created_date=now,
            created_by=current_user.username
        )
        db.add(customer_report)

        # Авто-скрининг AML: компания + директор + UBO
        try:
            from app.services.aml_auto_screening import auto_screen_on_kyc_approval
            await auto_screen_on_kyc_approval(
                client_id=client_id,
                db=db,
                triggered_by=current_user.user_id,
            )
        except Exception as e:
            logger.error(f"AML авто-скрининг не удался для {client_id}: {e}")

    await db.commit()
    await db.refresh(profile)
    
    status_messages = {
        KYCStatus.APPROVED.value: "KYC профиль одобрен",
        KYCStatus.REJECTED.value: "KYC профиль отклонен",
        KYCStatus.NEEDS_FIX.value: "KYC возвращен на доработку"
    }
    
    return KYCDecisionResponse(
        profile_id=profile.profile_id,
        client_id=profile.client_id,
        status=profile.status,
        decided_at=profile.decided_at,
        decided_by=profile.decided_by,
        decision_comment=profile.decision_comment,
        message=status_messages.get(decision.status, "Решение принято")
    )


# ======================================================================
# KYC EXCEL GENERATION
# ======================================================================

async def generate_kyc_excel(
    client_id: str,
    db: AsyncSession,
) -> BytesIO:
    result = await db.execute(
        select(Client).where(Client.client_id == client_id)
    )
    client = result.scalar_one_or_none()

    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    result = await db.execute(
        select(OnboardingKycProfile).where(OnboardingKycProfile.client_id == client_id)
    )
    profile = result.scalar_one_or_none()

    if not profile:
        raise HTTPException(status_code=404, detail="KYC profile not found")

    payload = profile.payload or {}
    corporate = payload.get("corporate", {})
    banking = payload.get("banking", {})
    declaration = payload.get("declaration", {})

    result = await db.execute(
        select(OnboardingKycUbo).where(OnboardingKycUbo.profile_id == profile.profile_id)
    )
    ubos = result.scalars().all()

    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "template",
        "KYC.xlsx"
    )

    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="KYC template file not found")

    output = await asyncio.to_thread(
        _build_kyc_excel, template_path, corporate, banking, declaration, ubos
    )
    return output


def _build_kyc_excel(template_path, corporate, banking, declaration, ubos) -> BytesIO:
    wb = load_workbook(template_path)
    ws = wb.active

    # 1. Corporate Details
    ws['D8'] = corporate.get("company_name")
    ws['J8'] = corporate.get("trading_name")
    ws['D10'] = corporate.get("incorporation_date")
    ws['J10'] = corporate.get("incorporation_country")
    ws['D16'] = corporate.get("registered_address")
    ws['J18'] = corporate.get("tax_id")
    ws['D22'] = corporate.get("registration_number")
    ws['D24'] = corporate.get("telephone")
    ws['D26'] = corporate.get("website")

    # 4. Banking Details
    ws['D66'] = banking.get("principal_bankers")
    ws['D68'] = banking.get("swift_bic")
    ws['D70'] = banking.get("bank_branch_address")
    ws['D72'] = banking.get("bank_city_country")
    ws['D74'] = banking.get("bank_account_name")
    ws['D76'] = banking.get("bank_account_currency")
    ws['D78'] = banking.get("bank_account_number")
    ws['D80'] = banking.get("bank_manager_contact")

    # 5. Ownership information (all UBOs side by side)
    # Each UBO occupies 3 columns: D:F, G:I, J:L
    ubo_start_cols = [4, 7, 10]  # D=4, G=7, J=10
    thin_border = Border(
        left=Side(style='thin', color='FFBFBFBF'),
        right=Side(style='thin', color='FFBFBFBF'),
        top=Side(style='thin', color='FFBFBFBF'),
        bottom=Side(style='thin', color='FFBFBFBF'),
    )
    center_align = Alignment(horizontal='center')
    ubo_font = Font(name='Calibri', size=12, color='00000000')

    for idx, ubo in enumerate(ubos[:3]):
        start_col = ubo_start_cols[idx]
        end_col = start_col + 2
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)

        # Merge cells for this UBO (skip first — already merged in template)
        if idx > 0:
            for row in [84, 86, 88]:
                ws.merge_cells(f'{start_letter}{row}:{end_letter}{row}')

        # Apply styling to every cell in the merged range
        for row in [84, 86, 88]:
            for col in range(start_col, end_col + 1):
                cl = get_column_letter(col)
                cell = ws[f'{cl}{row}']
                cell.border = thin_border
                cell.alignment = center_align
                cell.font = ubo_font

        # 5.1 Shareholder/UBO Name
        ws[f'{start_letter}84'] = ubo.ubo_name

        # 5.2 Shareholding %
        ws[f'{start_letter}86'] = float(ubo.shareholding_percent) if ubo.shareholding_percent else None

        # 5.3 Nationality and Country of Residence
        nationality_str = ""
        if ubo.nationality:
            nationality_str += ubo.nationality
        if ubo.residence_country:
            if nationality_str:
                nationality_str += " / "
            nationality_str += ubo.residence_country
        ws[f'{start_letter}88'] = nationality_str or None

    # 7. Signature
    ws['D95'] = declaration.get("authorized_person_name")
    ws['J95'] = declaration.get("signature_date")
    ws['D97'] = declaration.get("authorized_person_position")
    ws['J97'] = declaration.get("signature_location")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/clients/{client_id}/kyc/excel", summary="Сгенерировать KYC Excel из шаблона")
async def export_kyc_excel(
    client_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _check_client_access(db, client_id, current_user)

    excel_file = await generate_kyc_excel(client_id, db)

    filename = f"KYC_{client_id}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
