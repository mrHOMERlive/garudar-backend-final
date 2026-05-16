from datetime import datetime, date as date_type
from typing import Optional
import json
import re
import textwrap
import os
from decimal import Decimal
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import JSONResponse, Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from io import BytesIO
from openpyxl import load_workbook
from app.db import get_db
from app.models import User, OrderPobo, OrderStatusHistory, Client, Role, OrderPoboTerm, AuditLog, InstructionExport, InstructionExportItem, AppSetting, ExecutedOrder, PayeerAccount, TransactionReport, CustomerReport, OnboardingKycProfile, OrgDirectory
from app.schemas import OrderPoboDto, ErrorResponse, OrderPoboTermCreateUpdateRequest, OrderPoboTermDto, ExportInstructionRequest, ExportInstructionResponse, ExecutedOrderDto, ExecutedOrderUpdateRequest, LastInstructionExportResponse
from app.enums import KYCStatus
from app.deps import get_current_active_user

router = APIRouter(tags=["Payment Orders"])


def is_admin(user: User) -> bool:
    return user.role == Role.ADMIN.value


async def generate_order_id(db: AsyncSession, client: Client) -> tuple[str, int]:
    """
    Генерирует order_id в формате ORD-{client_numeric_id}-{order_sequence}
    Пример: клиент CL5 создает 10-й ордер → ORD-5-10
    
    Args:
        db: AsyncSession
        client: Client объект
        
    Returns:
        tuple[str, int]: (order_id, new_orders_count)
    """
    # Извлекаем числовую часть из client_id (например, CL5 → 5)
    client_numeric_id = client.client_id.replace("CL", "")
    
    # Увеличиваем счетчик ордеров для клиента
    new_orders_count = (client.orders_count or 0) + 1
    
    # Формируем order_id (убираем слеш для корректного URL роутинга)
    order_id = f"ORD-{client_numeric_id}-{new_orders_count}"
    
    return order_id, new_orders_count


def serialize_value(value):
    """Конвертирует типы данных в JSON-сериализуемые значения"""
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date_type)):
        return value.isoformat()
    return value


async def log_audit(
    db: AsyncSession,
    entity: str,
    entity_id: str,
    action: str,
    old_value: Optional[dict],
    new_value: Optional[dict],
    user_id: str,
):
    """Создает запись в audit_log для отслеживания изменений"""
    # Конвертируем Decimal, date, datetime в строки для JSON сериализации
    if old_value:
        old_value = {k: serialize_value(v) for k, v in old_value.items()}
    if new_value:
        new_value = {k: serialize_value(v) for k, v in new_value.items()}
    
    audit_entry = AuditLog(
        entity=entity,
        entity_id=entity_id,
        action=action,
        old_value=json.dumps(old_value, ensure_ascii=False) if old_value else None,
        new_value=json.dumps(new_value, ensure_ascii=False) if new_value else None,
        created_by=user_id,
        created_at=datetime.utcnow(),
    )
    db.add(audit_entry)
    return audit_entry


# ========== TXT Export Helper Functions ==========

def _clean(v) -> str:
    """Очистка значения от NaN и None"""
    if v is None or (isinstance(v, float) and str(v) == 'nan'):
        return ""
    s = str(v)
    return "" if s.lower() == "nan" else s.strip()


def _pad13(x: str) -> str:
    """Добавляет нули слева до 13 символов"""
    return _clean(x).zfill(13)


def _fmt_amt(v) -> str:
    """Форматирует сумму с 2 знаками после запятой"""
    s = _clean(v).replace(",", ".")
    try:
        return f"{float(s):.2f}"
    except:
        return "0.00"


def _wrap35(text: str) -> tuple[str, str, str]:
    """Разбивает текст на 3 строки по 35 символов"""
    s = " ".join(_clean(text).split())
    parts = textwrap.wrap(s, width=35, break_long_words=False, break_on_hyphens=False)
    return (parts + ["", "", ""])[:3]


def _split_remark_chunks(text: str, chunk_size: int = 30, parts: int = 4) -> list[str]:
    """Разбивает Remark на 4 части по 30 символов"""
    normalized = re.sub(r"\s+", " ", _clean(text))
    chunks = []
    for i in range(parts):
        start = i * chunk_size
        end = start + chunk_size
        piece = normalized[start:end]
        chunks.append(piece.strip() if piece else "")
    return chunks


def _pad_bic(b: str) -> str:
    """Дополняет BIC до 11 символов символом X"""
    b = _clean(b).upper()
    return b + "X" * (11 - len(b)) if b and len(b) < 11 else b


async def get_app_settings(db: AsyncSession) -> dict:
    """Получает настройки приложения из app_settings"""
    result = await db.execute(select(AppSetting))
    settings = result.scalars().all()
    return {s.key: s.value for s in settings}


@router.get(
    "/pobo",
    response_model=list[OrderPoboDto],
    summary="Получить все платежные поручения POBO",
)
async def get_orders(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    client_id: Optional[str] = Query(None, description="Фильтр по client_id (только для админа)"),
    include_deleted: bool = Query(False, description="Включить удаленные заказы"),
    skip: int = Query(0, ge=0, description="Пропустить N записей"),
    limit: int = Query(100, ge=1, le=500, description="Макс. кол-во записей"),
):
    if is_admin(current_user):
        query = select(OrderPobo)
        if not include_deleted:
            query = query.where(OrderPobo.deleted == False)
        if client_id:
            query = query.where(OrderPobo.client_id == client_id)
    else:
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None:
            return []

        query = select(OrderPobo).where(OrderPobo.client_id == client.client_id)
        if not include_deleted:
            query = query.where(OrderPobo.deleted == False)

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    orders = result.scalars().all()
    return [OrderPoboDto.model_validate(o, from_attributes=True) for o in orders]


@router.post(
    "/pobo",
    response_model=OrderPoboDto,
    summary="Создать новое платежное поручение POBO",
)
async def create_order(
    dto: OrderPoboDto,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    import logging
    
    logger = logging.getLogger(__name__)
    
    if dto.clientId:
        logger.info(f"Validating clientId from DTO: {dto.clientId}")
        result = await db.execute(select(Client).where(Client.client_id == dto.clientId))
        client = result.scalar_one_or_none()
        if client is None:
            logger.error(f"Invalid clientId in DTO: {dto.clientId} does not exist in clients table")
            return JSONResponse(status_code=400, content={"error": f"Invalid client_id: {dto.clientId}"})
        logger.info(f"Using validated clientId: {client.client_id}")
    else:
        logger.info(f"Looking for client by user_id: {current_user.user_id}")
        result = await db.execute(select(Client).where(Client.user_id == current_user.user_id))
        client = result.scalar_one_or_none()
        if client is None:
            logger.error(f"Client not found for user_id: {current_user.user_id}")
            return JSONResponse(status_code=400, content={"error": "Client not found for current user"})
        logger.info(f"Found client_id: {client.client_id}")

    # KYC-gate (ТЗ Sec 10.4): клиент в роли USER может создавать заявки
    # только при kyc_status='approved' либо kyc_override=True (флаг
    # выставляется администратором вручную для исключений).
    # Staff/Admin при создании от имени клиента проходят без проверки.
    if current_user.role == Role.USER.value:
        if client.kyc_status != KYCStatus.APPROVED.value and not client.kyc_override:
            logger.warning(
                f"Order creation blocked: client {client.client_id} kyc_status="
                f"{client.kyc_status}, kyc_override={client.kyc_override}"
            )
            # Compliance-audit: фиксируем заблокированную попытку.
            # Коммитим в отдельной транзакции, чтобы запись осталась даже
            # после raise HTTPException (FastAPI не откатывает то, что уже
            # commit'нуто).
            db.add(AuditLog(
                entity="clients",
                entity_id=client.client_id,
                action="KYC_GATE_DENIED",
                old_value=None,
                new_value=json.dumps({
                    "endpoint": "POST /orders/pobo",
                    "kyc_status": client.kyc_status,
                    "kyc_override": client.kyc_override,
                    "reason": "kyc_not_approved",
                    "amount": str(dto.amount) if dto.amount is not None else None,
                    "currency": dto.currency,
                }, ensure_ascii=False),
                created_by=current_user.username,
                created_at=datetime.utcnow(),
            ))
            await db.commit()
            raise HTTPException(
                status_code=403,
                detail={
                    "error": "kyc_not_approved",
                    "message": "KYC must be approved before creating orders",
                    "kyc_status": client.kyc_status,
                },
            )

    # Аккаунт на hold — заявки тоже запрещены (синхронизируем с UI-проверкой).
    if client.account_status == "hold":
        logger.warning(f"Order creation blocked: client {client.client_id} on hold")
        db.add(AuditLog(
            entity="clients",
            entity_id=client.client_id,
            action="ACCOUNT_HOLD_DENIED",
            old_value=None,
            new_value=json.dumps({
                "endpoint": "POST /orders/pobo",
                "account_status": client.account_status,
                "account_hold_reason": client.account_hold_reason,
                "amount": str(dto.amount) if dto.amount is not None else None,
                "currency": dto.currency,
            }, ensure_ascii=False),
            created_by=current_user.username,
            created_at=datetime.utcnow(),
        ))
        await db.commit()
        raise HTTPException(
            status_code=403,
            detail={
                "error": "account_on_hold",
                "message": "Account is on hold",
                "reason": client.account_hold_reason,
            },
        )

    # ТЗ Sec 5.2.2: bank_name / bank_address — серверная подстановка из
    # org_directory, чтобы клиент не мог через прямой API-вызов отправить
    # произвольные реквизиты для валидного BIC. UI делает то же самое для
    # UX, но финальное слово за бэком.
    final_bank_name = dto.bankName
    final_bank_address = dto.bankAddress
    bank_override_used = False  # для отложенного audit-лога после flush

    if dto.bankBic:
        if not dto.bankManualOverride:
            org = await db.scalar(
                select(OrgDirectory).where(
                    OrgDirectory.bic_swift_cd == dto.bankBic,
                    OrgDirectory.is_inactive.is_(False),
                    OrgDirectory.is_delete.is_(False),
                ).limit(1)
            )
            if org is None:
                logger.warning(
                    f"BIC {dto.bankBic} not found in org_directory; client "
                    f"{client.client_id} did not enable manual override"
                )
                raise HTTPException(
                    status_code=422,
                    detail={
                        "error": "bic_not_in_directory",
                        "message": (
                            "BIC not found in directory. Enable bank_manual_override "
                            "to enter bank details manually, or use a different BIC."
                        ),
                        "bic": dto.bankBic,
                    },
                )
            # Подменяем тем, что в директории — клиентский input игнорируется.
            final_bank_name = org.nm
            final_bank_address = ", ".join(filter(None, [
                org.addr_1, org.addr_2, org.addr_3, org.city_nm,
            ])) or None
        else:
            # Override включён — клиент сам отвечает за реквизиты.
            # Требуем не-пустой bank_name (иначе на TXT-инструкции уйдёт мусор).
            if not final_bank_name:
                raise HTTPException(
                    status_code=422,
                    detail={
                        "error": "bank_name_required_for_override",
                        "message": (
                            "bank_name is required when bank_manual_override=True"
                        ),
                    },
                )
            bank_override_used = True

    # Генерируем order_id в формате ORD/{client_numeric_id}-{order_sequence}
    order_id, new_orders_count = await generate_order_id(db, client)
    logger.info(f"Generated order_id: {order_id}")

    # Обновляем счетчик ордеров клиента
    client.orders_count = new_orders_count

    order = OrderPobo(
        order_id=order_id,
        client_id=client.client_id,
        amount=dto.amount,
        currency=dto.currency,
        counterparty_id=dto.counterpartyId,
        beneficiary_name=dto.beneficiaryName,
        beneficiary_adress=dto.beneficiaryAdress,
        destination_account=dto.destinationAccount,
        beneficiary_country=dto.beneficiaryCountry,
        bank_country=dto.bankCountry,
        bank_bic=dto.bankBic,
        bank_name=final_bank_name,
        bank_address=final_bank_address,
        bank_manual_override=dto.bankManualOverride,
        remark=dto.remark,
        invocie_required=dto.invocieRequired,
        invocie_received=dto.invocieReceived,
        payment_proof=dto.paymentProof,
        non_mandiri_execution=dto.nonMandiriExecution,
        invoice_number=dto.invoiceNumber,
        status="created",
        include=True,
        deleted=False,
        executed=False,
        created_at=datetime.utcnow(),
    )
    db.add(order)
    # Flush order first so FK-dependent rows can reference it
    await db.flush()

    # ТЗ Sec 5.2.2: фиксируем каждый override-кейс в audit_log.
    # Делаем это после flush, чтобы entity_id мог ссылаться на order_id.
    if bank_override_used:
        db.add(AuditLog(
            entity="orders_pobo",
            entity_id=order.order_id,
            action="BANK_OVERRIDE_USED",
            old_value=None,
            new_value=json.dumps({
                "client_id": client.client_id,
                "bank_bic": dto.bankBic,
                "bank_country": dto.bankCountry,
                "bank_name": final_bank_name,
                "bank_address": final_bank_address,
            }, ensure_ascii=False),
            created_by=current_user.username,
            created_at=datetime.utcnow(),
        ))

    # Создаем order_pobo_terms автоматически
    order_term = OrderPoboTerm(
        order_id=order.order_id,
        client_id=client.client_id,
        amount=dto.amount,
        currency=dto.currency,
        client_payment_currency=dto.clientPaymentCurrency,
    )
    db.add(order_term)

    status_history = OrderStatusHistory(
        order_id=order.order_id,
        old_status=None,
        new_status="created",
        changed_by=current_user.user_id,
        changed_at=datetime.utcnow(),
        comment="Created by client",
    )
    db.add(status_history)

    await db.commit()
    await db.refresh(order)
    return OrderPoboDto.model_validate(order, from_attributes=True)


@router.get(
    "/pobo/{order_id}",
    response_model=OrderPoboDto,
    summary="Получить платежное поручение POBO по ID",
)
async def get_order_by_id(
    order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(OrderPobo).where(OrderPobo.order_id == order_id))
    order = result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})
    
    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
    
    return OrderPoboDto.model_validate(order, from_attributes=True)


@router.put(
    "/pobo/{order_id}",
    response_model=OrderPoboDto,
    summary="Обновить платежное поручение POBO",
)
async def update_order(
    order_id: str,
    dto: OrderPoboDto,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(OrderPobo).where(OrderPobo.order_id == order_id))
    order = result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})
    
    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
    
    if dto.amount is not None:
        order.amount = dto.amount
    if dto.currency is not None:
        order.currency = dto.currency
    if dto.counterpartyId is not None:
        order.counterparty_id = dto.counterpartyId
    if dto.beneficiaryName is not None:
        order.beneficiary_name = dto.beneficiaryName
    if dto.beneficiaryAdress is not None:
        order.beneficiary_adress = dto.beneficiaryAdress
    if dto.destinationAccount is not None:
        order.destination_account = dto.destinationAccount
    if dto.beneficiaryCountry is not None:
        order.beneficiary_country = dto.beneficiaryCountry
    if dto.bankCountry is not None:
        order.bank_country = dto.bankCountry
    if dto.bankBic is not None:
        order.bank_bic = dto.bankBic
    if dto.bankName is not None:
        order.bank_name = dto.bankName
    if dto.bankAddress is not None:
        order.bank_address = dto.bankAddress
    if dto.remark is not None:
        order.remark = dto.remark
    if dto.invocieRequired is not None:
        order.invocie_required = dto.invocieRequired
    if dto.invocieReceived is not None:
        order.invocie_received = dto.invocieReceived
    if dto.paymentProof is not None:
        order.payment_proof = dto.paymentProof
    if dto.nonMandiriExecution is not None:
        order.non_mandiri_execution = dto.nonMandiriExecution
    if dto.invoiceNumber is not None:
        order.invoice_number = dto.invoiceNumber
    
    original_status = order.status  # сохраняем до мутации для проверки executed
    if dto.status is not None and dto.status != order.status:
        old_status = order.status
        order.status = dto.status
        order.last_status = datetime.utcnow()
        
        status_history = OrderStatusHistory(
            order_id=order.order_id,
            old_status=old_status,
            new_status=dto.status,
            changed_by=current_user.user_id,
            changed_at=datetime.utcnow(),
            comment=f"Status changed by {'admin' if is_admin(current_user) else 'client'}",
        )
        db.add(status_history)
    
    if dto.include is not None:
        order.include = dto.include
    if dto.deleted is not None:
        order.deleted = dto.deleted
    
    # Проверяем условия для создания ExecutedOrder
    should_create_executed_order = False
    if dto.executed is not None and dto.executed and not order.executed:
        # Проверяем что статус = 'released'
        if original_status == 'released' or (dto.status is not None and dto.status == 'released'):
            should_create_executed_order = True
        order.executed = dto.executed
    # Авто-executed при переходе статуса в 'released'
    elif dto.status == 'released' and original_status != 'released' and not order.executed:
        should_create_executed_order = True
        order.executed = True
    
    # Создаем запись в ExecutedOrder если необходимо
    if should_create_executed_order:
        # Проверяем, не существует ли уже запись для этого ордера
        existing_executed = await db.execute(
            select(ExecutedOrder).where(ExecutedOrder.source_order_id == order.order_id)
        )
        if existing_executed.scalar_one_or_none() is None:
            # Создаем новую запись в ExecutedOrder
            executed_order = ExecutedOrder(
                source_order_id=order.order_id,
                moved_at=datetime.utcnow(),
                moved_by=current_user.user_id,
            )
            db.add(executed_order)
            await db.flush()  # Получаем executed_id
            
            # Логируем создание ExecutedOrder в audit log
            await log_audit(
                db=db,
                entity="executed_orders",
                entity_id=str(executed_order.executed_id),
                action="INSERT",
                old_value=None,
                new_value={
                    "source_order_id": order.order_id,
                    "moved_at": executed_order.moved_at,
                    "moved_by": executed_order.moved_by,
                },
                user_id=current_user.username,
            )

            # Автоматически создаем TransactionReport при пометке ордера как Executed
            terms_result = await db.execute(
                select(OrderPoboTerm).where(OrderPoboTerm.order_id == order.order_id)
            )
            term = terms_result.scalar_one_or_none()

            transaction_id = f"TRX-{order.order_id}"
            existing_report = await db.execute(
                select(TransactionReport).where(TransactionReport.transaction_id == transaction_id)
            )
            if not existing_report.scalar_one_or_none():
                # Найти связанный CustomerReport и KYC-данные через Client
                customer_report_id = None
                sender_name = None
                sender_address = None
                sender_bank_bic = None
                sender_bank_name = None
                account_holder_name = None
                account_number = None

                if order.client_id:
                    client_result = await db.execute(
                        select(Client).where(Client.client_id == order.client_id)
                    )
                    client = client_result.scalar_one_or_none()
                    if client:
                        sender_name = client.client_name

                        # Найти CustomerReport по registration_number
                        if client.client_reg_number:
                            cr_result = await db.execute(
                                select(CustomerReport).where(
                                    CustomerReport.registration_number == client.client_reg_number
                                )
                            )
                            cr = cr_result.scalars().first()
                            if cr:
                                customer_report_id = cr.id

                        # Подтянуть banking-данные из KYC payload
                        kyc_result = await db.execute(
                            select(OnboardingKycProfile).where(
                                OnboardingKycProfile.client_id == order.client_id
                            )
                        )
                        kyc_profile = kyc_result.scalars().first()
                        if kyc_profile and kyc_profile.payload:
                            payload = kyc_profile.payload if isinstance(kyc_profile.payload, dict) else {}
                            corporate = payload.get("corporate", {})
                            banking = payload.get("banking", {})
                            sender_address = corporate.get("registered_address")
                            sender_bank_bic = banking.get("swift_bic")
                            sender_bank_name = banking.get("principal_bankers")
                            account_holder_name = banking.get("bank_account_name")
                            account_number = banking.get("bank_account_number")

                transaction_report = TransactionReport(
                    transaction_id=transaction_id,
                    date=date_type.today(),
                    customer_report_id=customer_report_id,
                    sender_name=sender_name,
                    sender_address=sender_address,
                    sender_bank_bic=sender_bank_bic,
                    sender_bank_name=sender_bank_name,
                    recipient_name=order.beneficiary_name,
                    recipient_address=order.beneficiary_adress,
                    currency=order.currency,
                    amount=order.amount,
                    transaction_purpose=order.remark,
                    transfer_fee=term.amount_remuneration if term else None,
                    account_holder_name=account_holder_name,
                    account_number=account_number,
                    transaction_type="outgoing",
                    transaction_method="SWIFT",
                    beneficiary_type="Corporate",
                    created_date=datetime.utcnow(),
                    created_by=current_user.username
                )
                db.add(transaction_report)

                # Автоматически создаём CustomerReport для recipient (counterparty)
                if order.beneficiary_name:
                    existing_counterparty = await db.execute(
                        select(CustomerReport).where(
                            CustomerReport.customer_type == "counterparty",
                            CustomerReport.name == order.beneficiary_name
                        )
                    )
                    if not existing_counterparty.scalar_one_or_none():
                        counterparty_report = CustomerReport(
                            customer_type="counterparty",
                            name=order.beneficiary_name,
                            address=order.beneficiary_adress,
                            recipient_name=order.beneficiary_name,
                            recipient_address=order.beneficiary_adress,
                            created_date=datetime.utcnow(),
                            created_by=current_user.username
                        )
                        db.add(counterparty_report)

    order.updated_at = datetime.utcnow()
    await db.commit()
    await db.refresh(order)
    return OrderPoboDto.model_validate(order, from_attributes=True)


@router.delete(
    "/pobo/{order_id}",
    status_code=204,
    summary="Удалить платежное поручение POBO",
)
async def delete_order(
    order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(OrderPobo).where(OrderPobo.order_id == order_id))
    order = result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})

    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})

    order.deleted = True
    order.updated_at = datetime.utcnow()
    await db.commit()
    return Response(status_code=204)


@router.post(
    "/pobo/{order_id}/terms",
    response_model=OrderPoboTermDto,
    summary="Создать или обновить условия для платежного поручения POBO",
)
async def create_or_update_order_terms(
    order_id: str,
    dto: OrderPoboTermCreateUpdateRequest,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Создает или обновляет условия для POBO заказа.
    
    Логика:
    - Если order_pobo_terms не существует → INSERT
    - Если существует → UPDATE
    - amount и currency копируются из orders_pobo
    - amount_remuneration и amount_to_be_paid вычисляются автоматически
    - exchange_rate_manual всегда = true
    """
    if not is_admin(current_user):
        return JSONResponse(status_code=403, content={"error": "Only admins can manage order terms"})
    
    # Валидация типа вознаграждения
    if dto.remuneration_type not in ["percent", "fixed"]:
        return JSONResponse(
            status_code=400,
            content={"error": "remuneration_type must be 'percent' or 'fixed'"}
        )
    
    # Валидация соответствия полей типу
    if dto.remuneration_type == "percent" and dto.remuneration_percentage is None:
        return JSONResponse(
            status_code=400,
            content={"error": "remuneration_percentage is required when type is 'percent'"}
        )
    
    if dto.remuneration_type == "fixed" and dto.remuneration_fixed is None:
        return JSONResponse(
            status_code=400,
            content={"error": "remuneration_fixed is required when type is 'fixed'"}
        )
    
    # Получаем заказ POBO
    order_result = await db.execute(select(OrderPobo).where(OrderPobo.order_id == order_id))
    order = order_result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})
    
    # Запрещаем изменение terms для выполненных заказов
    if order.executed:
        return JSONResponse(
            status_code=400,
            content={"error": "Cannot modify terms for executed orders"}
        )
    
    # Проверяем существование terms
    term_result = await db.execute(
        select(OrderPoboTerm).where(OrderPoboTerm.order_id == order_id)
    )
    existing_term = term_result.scalar_one_or_none()
    
    # Вычисляем amount_remuneration
    if dto.remuneration_type == "percent":
        amount_remuneration = (order.amount or 0) * (dto.remuneration_percentage or 0) / 100
    else:
        amount_remuneration = dto.remuneration_fixed or 0
    
    # Вычисляем amount_to_be_paid
    amount_to_be_paid = (order.amount or 0) + amount_remuneration
    
    if existing_term:
        # UPDATE существующей записи
        # Сохраняем старые значения для аудита
        old_values = {
            "remuneration_type": existing_term.remuneration_type,
            "remuneration_percentage": existing_term.remuneration_percentage,
            "remuneration_fixed": existing_term.remuneration_fixed,
            "exchange_rate": existing_term.exchange_rate,
            "client_payment_currency": existing_term.client_payment_currency,
            "date_paid": existing_term.date_paid,
            "data_fixing": existing_term.data_fixing,
            "amount_remuneration": existing_term.amount_remuneration,
            "amount_to_be_paid": existing_term.amount_to_be_paid,
            "GAN_bank_name": existing_term.GAN_bank_name,
            "GAN_bank_account": existing_term.GAN_bank_account,
            "date_report": existing_term.date_report,
            "conversion_method": existing_term.conversion_method,
            "base_currency": existing_term.base_currency,
            "FX": existing_term.FX,
            "executing_bank": existing_term.executing_bank,
            "FX_executing_bank": existing_term.FX_executing_bank,
            "status": existing_term.status,
            "bank_statement_in_type": existing_term.bank_statement_in_type,
            "bank_statement_in_id": existing_term.bank_statement_in_id,
            "bank_statement_out_type": existing_term.bank_statement_out_type,
            "bank_statement_out_id": existing_term.bank_statement_out_id,
            "amount_to_be_paid_target_cur": existing_term.amount_to_be_paid_target_cur,
            "amount_paid_target_cur": existing_term.amount_paid_target_cur,
            "doc_paid_no": existing_term.doc_paid_no,
            "doc_paid_date": existing_term.doc_paid_date,
            "payment_proof_no": existing_term.payment_proof_no,
            "payment_proof_date": existing_term.payment_proof_date,
            "description": existing_term.description,
        }
        
        existing_term.remuneration_type = dto.remuneration_type
        existing_term.remuneration_percentage = dto.remuneration_percentage
        existing_term.remuneration_fixed = dto.remuneration_fixed
        existing_term.exchange_rate = dto.exchange_rate
        existing_term.exchange_rate_manual = dto.exchange_rate  # Сохраняем как manual
        existing_term.client_payment_currency = dto.client_payment_currency
        existing_term.date_paid = dto.date_paid
        existing_term.data_fixing = dto.data_fixing
        existing_term.amount_remuneration = amount_remuneration
        existing_term.amount_to_be_paid = amount_to_be_paid
        # Обновляем amount и currency из orders_pobo
        existing_term.amount = order.amount
        existing_term.currency = order.currency
        # Обновляем новые поля
        existing_term.GAN_bank_name = dto.GAN_bank_name
        existing_term.GAN_bank_account = dto.GAN_bank_account
        existing_term.date_report = dto.date_report
        existing_term.conversion_method = dto.conversion_method
        # Конвертируем пустые строки в None для полей с FK
        existing_term.base_currency = dto.base_currency or None
        existing_term.FX = dto.FX
        existing_term.executing_bank = dto.executing_bank or None
        existing_term.FX_executing_bank = dto.FX_executing_bank
        existing_term.status = dto.status
        existing_term.bank_statement_in_type = dto.bank_statement_in_type
        existing_term.bank_statement_in_id = dto.bank_statement_in_id
        existing_term.bank_statement_out_type = dto.bank_statement_out_type
        existing_term.bank_statement_out_id = dto.bank_statement_out_id
        existing_term.amount_to_be_paid_target_cur = dto.amount_to_be_paid_target_cur
        existing_term.amount_paid_target_cur = dto.amount_paid_target_cur
        existing_term.doc_paid_no = dto.doc_paid_no
        existing_term.doc_paid_date = dto.doc_paid_date
        existing_term.payment_proof_no = dto.payment_proof_no
        existing_term.payment_proof_date = dto.payment_proof_date
        existing_term.date_paid = dto.payment_proof_date
        existing_term.description = dto.description
        
        # Новые значения для аудита
        new_values = {
            "remuneration_type": dto.remuneration_type,
            "remuneration_percentage": dto.remuneration_percentage,
            "remuneration_fixed": dto.remuneration_fixed,
            "exchange_rate": dto.exchange_rate,
            "client_payment_currency": dto.client_payment_currency,
            "date_paid": dto.date_paid,
            "data_fixing": dto.data_fixing,
            "amount_remuneration": amount_remuneration,
            "amount_to_be_paid": amount_to_be_paid,
            "GAN_bank_name": dto.GAN_bank_name,
            "GAN_bank_account": dto.GAN_bank_account,
            "date_report": dto.date_report,
            "conversion_method": dto.conversion_method,
            "base_currency": dto.base_currency,
            "FX": dto.FX,
            "executing_bank": dto.executing_bank,
            "FX_executing_bank": dto.FX_executing_bank,
            "status": dto.status,
            "bank_statement_in_type": dto.bank_statement_in_type,
            "bank_statement_in_id": dto.bank_statement_in_id,
            "bank_statement_out_type": dto.bank_statement_out_type,
            "bank_statement_out_id": dto.bank_statement_out_id,
            "amount_to_be_paid_target_cur": dto.amount_to_be_paid_target_cur,
            "amount_paid_target_cur": dto.amount_paid_target_cur,
            "doc_paid_no": dto.doc_paid_no,
            "doc_paid_date": dto.doc_paid_date,
            "payment_proof_no": dto.payment_proof_no,
            "payment_proof_date": dto.payment_proof_date,
            "description": dto.description,
        }
        
        # Логируем изменение
        await log_audit(
            db=db,
            entity="order_pobo_terms",
            entity_id=str(existing_term.term_id),
            action="UPDATE",
            old_value=old_values,
            new_value=new_values,
            user_id=current_user.username,
        )
        
        await db.commit()
        await db.refresh(existing_term)
        return OrderPoboTermDto.model_validate(existing_term, from_attributes=True)
    else:
        # INSERT новой записи
        new_term = OrderPoboTerm(
            order_id=order_id,
            client_id=order.client_id,
            amount=order.amount,
            currency=order.currency,
            client_payment_currency=dto.client_payment_currency,
            date_paid=dto.payment_proof_date,
            data_fixing=dto.data_fixing,
            remuneration_type=dto.remuneration_type,
            remuneration_percentage=dto.remuneration_percentage,
            remuneration_fixed=dto.remuneration_fixed,
            exchange_rate=dto.exchange_rate,
            exchange_rate_manual=dto.exchange_rate,  # Всегда true (сохраняем значение)
            amount_remuneration=amount_remuneration,
            amount_to_be_paid=amount_to_be_paid,
            GAN_bank_name=dto.GAN_bank_name,
            GAN_bank_account=dto.GAN_bank_account,
            date_report=dto.date_report,
            conversion_method=dto.conversion_method,
            base_currency=dto.base_currency or None,  # Конвертируем пустые строки в None для полей с FK
            FX=dto.FX,
            executing_bank=dto.executing_bank or None,
            FX_executing_bank=dto.FX_executing_bank,
            status=dto.status,
            bank_statement_in_type=dto.bank_statement_in_type,
            bank_statement_in_id=dto.bank_statement_in_id,
            bank_statement_out_type=dto.bank_statement_out_type,
            bank_statement_out_id=dto.bank_statement_out_id,
            amount_to_be_paid_target_cur=dto.amount_to_be_paid_target_cur,
            amount_paid_target_cur=dto.amount_paid_target_cur,
            doc_paid_no=dto.doc_paid_no,
            doc_paid_date=dto.doc_paid_date,
            payment_proof_no=dto.payment_proof_no,
            payment_proof_date=dto.payment_proof_date,
            description=dto.description,
        )
        db.add(new_term)
        await db.flush()  # Получаем term_id перед commit
        
        # Логируем создание
        new_values = {
            "order_id": order_id,
            "client_id": order.client_id,
            "remuneration_type": dto.remuneration_type,
            "remuneration_percentage": dto.remuneration_percentage,
            "remuneration_fixed": dto.remuneration_fixed,
            "exchange_rate": dto.exchange_rate,
            "client_payment_currency": dto.client_payment_currency,
            "date_paid": dto.date_paid,
            "data_fixing": dto.data_fixing,
            "amount_remuneration": amount_remuneration,
            "amount_to_be_paid": amount_to_be_paid,
            "GAN_bank_name": dto.GAN_bank_name,
            "GAN_bank_account": dto.GAN_bank_account,
            "date_report": dto.date_report,
            "conversion_method": dto.conversion_method,
            "base_currency": dto.base_currency,
            "FX": dto.FX,
            "executing_bank": dto.executing_bank,
            "FX_executing_bank": dto.FX_executing_bank,
            "status": dto.status,
            "bank_statement_in_type": dto.bank_statement_in_type,
            "bank_statement_in_id": dto.bank_statement_in_id,
            "bank_statement_out_type": dto.bank_statement_out_type,
            "bank_statement_out_id": dto.bank_statement_out_id,
            "amount_to_be_paid_target_cur": dto.amount_to_be_paid_target_cur,
            "amount_paid_target_cur": dto.amount_paid_target_cur,
            "doc_paid_no": dto.doc_paid_no,
            "doc_paid_date": dto.doc_paid_date,
            "payment_proof_no": dto.payment_proof_no,
            "payment_proof_date": dto.payment_proof_date,
            "description": dto.description,
        }
        
        await log_audit(
            db=db,
            entity="order_pobo_terms",
            entity_id=str(new_term.term_id),
            action="INSERT",
            old_value=None,
            new_value=new_values,
            user_id=current_user.username,
        )
        
        await db.commit()
        await db.refresh(new_term)
        return OrderPoboTermDto.model_validate(new_term, from_attributes=True)


@router.get(
    "/pobo/{order_id}/terms",
    response_model=OrderPoboTermDto,
    summary="Получить условия для платежного поручения POBO",
)
async def get_order_terms(
    order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Получает условия для POBO заказа.
    """
    # Проверяем доступ к заказу
    order_result = await db.execute(select(OrderPobo).where(OrderPobo.order_id == order_id))
    order = order_result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})
    
    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
    
    # Получаем terms
    term_result = await db.execute(
        select(OrderPoboTerm).where(OrderPoboTerm.order_id == order_id)
    )
    term = term_result.scalar_one_or_none()
    if term is None:
        return JSONResponse(status_code=404, content={"error": "Order terms not found"})
    
    return OrderPoboTermDto.model_validate(term, from_attributes=True)


@router.post(
    "/pobo/export-txt",
    summary="Экспорт TXT инструкций для Bank Mandiri (73 поля)",
)
async def export_txt_instructions(
    request: ExportInstructionRequest,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Генерирует TXT файл с инструкциями для Bank Mandiri.
    
    - Формирует строки из 73 полей для каждого заказа
    - Исключает заказы с non_mandiri_execution=true
    - Логирует в instruction_exports и instruction_export_items
    - Обновляет статус orders_pobo
    - Возвращает файл <YYYYMMDD>_instruction.txt для скачивания
    """
    if not is_admin(current_user):
        return JSONResponse(status_code=403, content={"error": "Only admins can export instructions"})
    
    # Получаем настройки приложения
    settings = await get_app_settings(db)
    sender_email = settings.get("sender_email", "garudainternasional7@gmail.com")
    static_code = settings.get("static_code", "2630")
    debit_account = settings.get("debit_account_no", "")
    ben_type_default = settings.get("beneficiary_type_default", "INU")
    
    # Текущая дата для имени файла
    tx_date = datetime.now().strftime("%Y%m%d")
    file_name = f"{tx_date}_instruction.txt"
    
    lines = []
    exported_orders = []
    skipped_orders = []
    
    # Обрабатываем каждый заказ
    for order_id in request.order_ids:
        # Получаем заказ с terms
        order_result = await db.execute(
            select(OrderPobo).where(OrderPobo.order_id == order_id)
        )
        order = order_result.scalar_one_or_none()
        
        if not order:
            skipped_orders.append({"order_id": order_id, "reason": "Order not found"})
            continue
        
        # Пропускаем заказы с non_mandiri_execution=true
        if order.non_mandiri_execution:
            skipped_orders.append({"order_id": order_id, "reason": "non_mandiri_execution is true"})
            continue
        
        # Получаем terms для заказа
        term_result = await db.execute(
            select(OrderPoboTerm).where(OrderPoboTerm.order_id == order_id)
        )
        term = term_result.scalar_one_or_none()
        
        if not term:
            skipped_orders.append({"order_id": order_id, "reason": "Terms not found"})
            continue
        
        # Формируем данные для 73 полей
        debit = _pad13(term.GAN_bank_account or "")
        dest = _clean(order.destination_account)
        name = _clean(order.beneficiary_name)
        
        # Адрес бенефициара (разбиваем на 3 строки по 35 символов)
        addr_full = _clean(order.beneficiary_adress or "")
        a1, a2, a3 = _wrap35(addr_full)
        
        # Валюта и сумма
        cur = _clean(order.currency).upper()
        amt_s = _fmt_amt(order.amount)
        
        # Remark и Invoice
        inv_remark = _clean(order.invoice_number or "")
        ref = _clean(order.order_id)
        bic = _pad_bic(order.bank_bic or "")
        remark = _clean(order.remark or "").upper()
        ben_t = _clean(ben_type_default).upper()
        invoice = _clean(order.invoice_number or "")
        
        # Разбиваем remark на 4 части по 30 символов
        rem_infos = _split_remark_chunks(remark)
        
        # Формируем массив из 73 полей
        f = [""] * 73
        
        # Позиции 1-3: M, 1, дата
        f[0:3] = ["M", "1", tx_date]
        
        # Позиция 8: Debit Account
        f[7] = debit
        
        # Позиция 9: Destination Account
        f[8] = dest
        
        # Позиция 10: Beneficiary Name
        f[9] = name
        
        # Позиции 11-13: Beneficiary Address (3 строки)
        f[10:13] = [a1, a2, a3]
        
        # Позиции 14-16: Currency, Amount, Currency
        f[13:16] = [cur, amt_s, cur]
        
        # Позиция 19: INV/REMARK (номер инвойса)
        f[18] = inv_remark
        
        # Позиция 20: Transaction Reference
        f[19] = ref
        
        # Позиция 21: Beneficiary Type (INU)
        f[20] = ben_t
        
        # Позиция 22: BIC
        f[21] = bic
        
        # Позиции 23-27: пустые (5 полей между BIC и Y)
        # f[22] через f[26] остаются пустыми
        
        # Позиция 28: Y
        f[27] = "Y"
        
        # Позиция 29: Email
        f[28] = "info@garudar.id"
        
        # Позиция 34: Static Code
        f[33] = static_code
        
        # Позиция 35: Full Remark
        f[34] = remark
        
        # Позиции 36-43: RFB + Remark chunks (4 пары)
        pairs = [
            (35, 36, rem_infos[0]),
            (37, 38, rem_infos[1]),
            (39, 40, rem_infos[2]),
            (41, 42, rem_infos[3])
        ]
        for rfb_idx, info_idx, value in pairs:
            if value:
                f[rfb_idx] = "RFB"
                f[info_idx] = value
        
        # Позиция 50: OUR (charge type)
        f[49] = "OUR"
        
        # Позиция 51: SPI
        f[50] = "SPI"
        
        # Позиция 52: Invoice
        f[51] = invoice
        
        # Позиции 72-73: E, EPD
        f[71] = "E"
        f[72] = "EPD"
        
        # Формируем строку с разделителями ;
        line = ";".join(f) + ";"
        lines.append(line)
        exported_orders.append(order_id)
    
    if not lines:
        return JSONResponse(
            status_code=400,
            content={"error": "No valid orders to export", "skipped": skipped_orders}
        )
    
    # Создаем запись в instruction_exports
    export_record = InstructionExport(
        export_date=datetime.now().date(),
        created_by=current_user.user_id,
        file_name=file_name,
        file_url="",
        export_params=json.dumps({
            "sender_email": sender_email,
            "static_code": static_code,
            "order_count": len(exported_orders)
        }, ensure_ascii=False),
        created_at=datetime.utcnow(),
    )
    db.add(export_record)
    await db.flush()
    
    # Создаем записи в instruction_export_items для каждого экспортированного заказа
    for order_id in exported_orders:
        export_item = InstructionExportItem(
            export_id=export_record.export_id,
            order_id=order_id,
            included=True,
        )
        db.add(export_item)
        
        # Обновляем статус заказа - добавляем запись в историю
        status_history = OrderStatusHistory(
            order_id=order_id,
            old_status=None,
            new_status="instruction_exported",
            changed_by=current_user.user_id,
            changed_at=datetime.utcnow(),
            comment=f"Exported to {file_name}",
        )
        db.add(status_history)
    
    await db.commit()
    
    # Формируем содержимое файла
    txt_content = "\n".join(lines)
    
    # Возвращаем файл для скачивания
    file_bytes = BytesIO(txt_content.encode("utf-8"))
    
    return StreamingResponse(
        file_bytes,
        media_type="text/plain",
        headers={
            "Content-Disposition": f"attachment; filename={file_name}"
        }
    )


async def generate_order_excel(
    order_id: str,
    db: AsyncSession,
) -> BytesIO:
    result = await db.execute(
        select(OrderPobo)
        .where(OrderPobo.order_id == order_id)
        .options(selectinload(OrderPobo.client), selectinload(OrderPobo.terms))
    )
    order = result.scalar_one_or_none()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    client = order.client
    terms = order.terms
    
    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "template",
        "order.xlsx"
    )
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template file not found")
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    if client:
        ws['B8'] = client.client_name
        ws['B9'] = client.doc_id
    
    ws['B6'] = order.order_id
    ws['B24'] = order.amount
    ws['B25'] = order.currency
    ws['B13'] = order.beneficiary_name
    ws['B14'] = order.beneficiary_adress
    ws['B15'] = order.beneficiary_country
    ws['B16'] = order.bank_name
    ws['B17'] = order.bank_bic
    ws['B18'] = order.bank_address
    ws['B19'] = order.bank_country
    ws['B20'] = order.destination_account
    ws['B49'] = order.remark
    ws['B7'] = order.created_at
    ws['B21'] = order.currency
    ws['B28'] = order.currency
    ws['B31'] = order.currency
    ws['B36'] = order.currency
    
    if terms:
        first_term = terms[0]
        ws['B34'] = first_term.client_payment_currency
        ws['B35'] = first_term.base_currency
        ws['B27'] = first_term.amount_remuneration
        ws['B30'] = first_term.amount_to_be_paid
        
        if first_term.conversion_method == 'central_bank':
            ws['B37'] = "☒ Central Bank Rate    ☐ Executing Bank Rate   ☐ Other (specify): __________"
        elif first_term.conversion_method == 'executing_bank':
            ws['B37'] = "☒ Executing Bank Rate   ☐ No Conversion"
        elif first_term.conversion_method == 'manual':
            ws['B37'] = "☐ Central Bank Rate    ☐ Executing Bank Rate   ☒ Other (specify): __________"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


@router.get(
    "/orders/{order_id}/excel",
    summary="Сгенерировать Excel ордер",
)
async def export_order_excel(
    order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    excel_file = await generate_order_excel(order_id, db)
    
    filename = f"order_{order_id}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# ========== ExecutedOrder Endpoints ==========

@router.get(
    "/executed-orders",
    response_model=list[ExecutedOrderDto],
    summary="Получить все выполненные заказы",
)
async def get_executed_orders(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
    skip: int = Query(0, ge=0, description="Пропустить N записей"),
    limit: int = Query(100, ge=1, le=500, description="Макс. кол-во записей"),
):
    """
    Получает все выполненные заказы.
    - Администраторы видят все выполненные заказы
    - Клиенты видят только свои выполненные заказы
    """
    if is_admin(current_user):
        query = select(ExecutedOrder)
    else:
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()

        if client is None:
            return []

        query = (
            select(ExecutedOrder)
            .join(OrderPobo, ExecutedOrder.source_order_id == OrderPobo.order_id)
            .where(OrderPobo.client_id == client.client_id)
        )

    query = query.offset(skip).limit(limit)
    result = await db.execute(query)
    executed_orders = result.scalars().all()
    return [ExecutedOrderDto.model_validate(eo, from_attributes=True) for eo in executed_orders]


@router.get(
    "/executed-orders/{source_order_id}",
    response_model=ExecutedOrderDto,
    summary="Получить выполненный заказ по source_order_id",
)
async def get_executed_order_by_id(
    source_order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Получает выполненный заказ по source_order_id.
    - Администраторы могут видеть любой выполненный заказ
    - Клиенты могут видеть только свои выполненные заказы
    """
    result = await db.execute(
        select(ExecutedOrder).where(ExecutedOrder.source_order_id == source_order_id)
    )
    executed_order = result.scalar_one_or_none()
    
    if executed_order is None:
        return JSONResponse(status_code=404, content={"error": "Executed order not found"})
    
    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        
        if client is None:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
        
        order_result = await db.execute(
            select(OrderPobo).where(OrderPobo.order_id == source_order_id)
        )
        order = order_result.scalar_one_or_none()
        
        if order is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
    
    return ExecutedOrderDto.model_validate(executed_order, from_attributes=True)


@router.put(
    "/executed-orders/{source_order_id}",
    response_model=ExecutedOrderDto,
    summary="Обновить выполненный заказ (только для админа)",
)
async def update_executed_order(
    source_order_id: str,
    dto: ExecutedOrderUpdateRequest,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Обновляет выполненный заказ по source_order_id. Доступно только для администраторов.
    Все изменения логируются в audit_log.
    """
    if not is_admin(current_user):
        return JSONResponse(status_code=403, content={"error": "Only admins can update executed orders"})
    
    result = await db.execute(
        select(ExecutedOrder).where(ExecutedOrder.source_order_id == source_order_id)
    )
    executed_order = result.scalar_one_or_none()
    
    if executed_order is None:
        return JSONResponse(status_code=404, content={"error": "Executed order not found"})
    
    # Сохраняем старые значения для audit log
    old_values = {
        "doc_package_status": executed_order.doc_package_status,
        "mt103_status": executed_order.mt103_status,
        "settled_status": executed_order.settled_status,
        "refund_flag": executed_order.refund_flag,
        "staff_description": executed_order.staff_description,
        "mt103_file_url": executed_order.mt103_file_url,
        "mt103_no": executed_order.mt103_no,
        "mt103_date": executed_order.mt103_date,
        "transaction_status_file_url": executed_order.transaction_status_file_url,
        "transaction_status_no": executed_order.transaction_status_no,
        "transaction_status_date": executed_order.transaction_status_date,
        "transaction_status_status": executed_order.transaction_status_status,
        "act_report_file_url": executed_order.act_report_file_url,
        "act_report_no": executed_order.act_report_no,
        "act_report_date": executed_order.act_report_date,
    }
    
    # Обновляем только переданные поля
    if dto.doc_package_status is not None:
        executed_order.doc_package_status = dto.doc_package_status
    if dto.mt103_status is not None:
        executed_order.mt103_status = dto.mt103_status
    if dto.settled_status is not None:
        executed_order.settled_status = dto.settled_status
    if dto.refund_flag is not None:
        executed_order.refund_flag = dto.refund_flag
    if dto.staff_description is not None:
        executed_order.staff_description = dto.staff_description
    if dto.mt103_file_url is not None:
        executed_order.mt103_file_url = dto.mt103_file_url
    if dto.mt103_no is not None:
        executed_order.mt103_no = dto.mt103_no
    if dto.mt103_date is not None:
        executed_order.mt103_date = dto.mt103_date
    if dto.transaction_status_file_url is not None:
        executed_order.transaction_status_file_url = dto.transaction_status_file_url
    if dto.transaction_status_no is not None:
        executed_order.transaction_status_no = dto.transaction_status_no
    if dto.transaction_status_date is not None:
        executed_order.transaction_status_date = dto.transaction_status_date
    if dto.transaction_status_status is not None:
        executed_order.transaction_status_status = dto.transaction_status_status
    if dto.act_report_file_url is not None:
        executed_order.act_report_file_url = dto.act_report_file_url
    if dto.act_report_no is not None:
        executed_order.act_report_no = dto.act_report_no
    if dto.act_report_date is not None:
        executed_order.act_report_date = dto.act_report_date
    
    await db.flush()
    
    # Новые значения для audit log
    new_values = {
        "doc_package_status": executed_order.doc_package_status,
        "mt103_status": executed_order.mt103_status,
        "settled_status": executed_order.settled_status,
        "refund_flag": executed_order.refund_flag,
        "staff_description": executed_order.staff_description,
        "mt103_file_url": executed_order.mt103_file_url,
        "mt103_no": executed_order.mt103_no,
        "mt103_date": executed_order.mt103_date,
        "transaction_status_file_url": executed_order.transaction_status_file_url,
        "transaction_status_no": executed_order.transaction_status_no,
        "transaction_status_date": executed_order.transaction_status_date,
        "transaction_status_status": executed_order.transaction_status_status,
        "act_report_file_url": executed_order.act_report_file_url,
        "act_report_no": executed_order.act_report_no,
        "act_report_date": executed_order.act_report_date,
    }
    
    # Логируем изменение в audit_log
    await log_audit(
        db=db,
        entity="executed_orders",
        entity_id=str(executed_order.executed_id),
        action="UPDATE",
        old_value=old_values,
        new_value=new_values,
        user_id=current_user.username,
    )
    
    await db.commit()
    await db.refresh(executed_order)
    return ExecutedOrderDto.model_validate(executed_order, from_attributes=True)


async def generate_act_report_excel(
    source_order_id: str,
    db: AsyncSession,
) -> BytesIO:
    executed_result = await db.execute(
        select(ExecutedOrder).where(ExecutedOrder.source_order_id == source_order_id)
    )
    executed_order = executed_result.scalar_one_or_none()
    
    if not executed_order:
        raise HTTPException(status_code=404, detail="Executed order not found")
    
    order_result = await db.execute(
        select(OrderPobo)
        .where(OrderPobo.order_id == source_order_id)
        .options(selectinload(OrderPobo.client), selectinload(OrderPobo.terms))
    )
    order = order_result.scalar_one_or_none()

    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    client = order.client
    terms = order.terms
    
    payeer_account = None
    if terms and terms[0].GAN_bank_account:
        payeer_result = await db.execute(
            select(PayeerAccount).where(PayeerAccount.account_no == terms[0].GAN_bank_account)
        )
        payeer_account = payeer_result.scalar_one_or_none()
    
    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "template",
        "act.xlsx"
    )
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Act template file not found")
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    ws['B6'] = order.order_id
    ws['B43'] = order.beneficiary_name
    ws['B44'] = order.bank_name
    ws['B45'] = order.destination_account
    
    if terms:
        first_term = terms[0]
        ws['B7'] = first_term.date_report
        ws['B12'] = first_term.date_paid
        ws['B13'] = first_term.client_payment_currency
        ws['B14'] = first_term.amount_to_be_paid_target_cur
        ws['B24'] = first_term.client_payment_currency
        ws['B25'] = first_term.base_currency
        ws['B26'] = first_term.conversion_method
        ws['B28'] = first_term.exchange_rate
        ws['B29'] = first_term.data_fixing
        ws['B30'] = first_term.amount_to_be_paid
        ws['B32'] = first_term.remuneration_percentage
        ws['B33'] = first_term.amount_remuneration
        ws['B42'] = first_term.base_currency
        ws['B47'] = first_term.date_report
        ws['B49'] = first_term.FX_executing_bank
        ws['B50'] = first_term.amount
        ws['B52'] = first_term.amount
        
        if first_term.executing_bank:
            executing_bank_result = await db.execute(
                select(PayeerAccount).where(PayeerAccount.account_no == first_term.executing_bank)
            )
            executing_bank_account = executing_bank_result.scalar_one_or_none()
            if executing_bank_account:
                ws['B48'] = executing_bank_account.bank_name

        
        
    
    if payeer_account:
        ws['B15'] = payeer_account.bank_name
        ws['B16'] = payeer_account.bank_address
        ws['B17'] = payeer_account.bank_bic
        ws['B19'] = payeer_account.account_no
        ws['B27'] = payeer_account.bank_country

    if client:
        ws['B8'] = client.client_name
        ws['A80'] = client.client_name
        ws['B9'] = client.doc_id
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


@router.get(
    "/executed-orders/{source_order_id}/act-report",
    summary="Сгенерировать Act Report Excel для выполненного заказа",
)
async def export_act_report_excel(
    source_order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    if not is_admin(current_user):
        return JSONResponse(status_code=403, content={"error": "Only admins can generate act reports"})
    
    excel_file = await generate_act_report_excel(source_order_id, db)
    
    filename = f"act_report_{source_order_id}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get(
    "/pobo/{order_id}/last-txt-export",
    response_model=LastInstructionExportResponse,
    summary="Получить дату последнего экспорта txt инструкции для заказа",
)
async def get_last_txt_export_date(
    order_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Возвращает информацию о последнем экспорте txt инструкции для заказа.
    
    - **order_id**: ID заказа POBO
    - **last_export_date**: Дата последнего экспорта (created_at из instruction_exports)
    - **export_count**: Количество раз, когда заказ был экспортирован
    """
    # Проверяем существование заказа
    order_result = await db.execute(
        select(OrderPobo).where(OrderPobo.order_id == order_id)
    )
    order = order_result.scalar_one_or_none()
    if order is None:
        return JSONResponse(status_code=404, content={"error": "Order not found"})
    
    # Проверяем доступ (клиенты могут видеть только свои заказы)
    if not is_admin(current_user):
        client_result = await db.execute(
            select(Client).where(Client.user_id == current_user.user_id)
        )
        client = client_result.scalar_one_or_none()
        if client is None or order.client_id != client.client_id:
            return JSONResponse(status_code=403, content={"error": "Access denied"})
    
    # Получаем все экспорты для данного order_id с JOIN к instruction_exports
    from sqlalchemy import func, desc
    
    query = (
        select(InstructionExport.created_at, InstructionExport.export_id)
        .join(InstructionExportItem, InstructionExport.export_id == InstructionExportItem.export_id)
        .where(InstructionExportItem.order_id == order_id)
        .where(InstructionExportItem.included == True)
        .order_by(desc(InstructionExport.created_at))
    )
    
    result = await db.execute(query)
    exports = result.all()
    
    export_count = len(exports)
    last_export_date = exports[0].created_at if exports else None
    
    if export_count == 0:
        message = "Заказ еще не был экспортирован в txt инструкцию"
    elif export_count == 1:
        message = f"Заказ был экспортирован 1 раз"
    else:
        message = f"Заказ был экспортирован {export_count} раз"
    
    return LastInstructionExportResponse(
        order_id=order_id,
        last_export_date=last_export_date,
        export_count=export_count,
        message=message
    )
