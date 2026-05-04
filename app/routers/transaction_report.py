import asyncio
import json
import os
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from app.db import get_db
from app.models import User, TransactionReport, AuditLog
from app.schemas import TransactionReportCreate, TransactionReportUpdate, TransactionReportDto
from app.deps import require_admin

router = APIRouter(tags=["Transaction Report"])


@router.get("/transaction-report", response_model=list[TransactionReportDto])
async def get_transaction_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Получить все отчеты по транзакциям"""
    result = await db.execute(
        select(TransactionReport).order_by(TransactionReport.date.desc())
    )
    reports = result.scalars().all()
    return [TransactionReportDto.model_validate(report, from_attributes=True) for report in reports]


@router.get("/transaction-report/export/excel", summary="Экспорт отчета по транзакциям в Excel")
async def export_transaction_report_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Выгрузка всех записей TransactionReport в Excel по шаблону"""
    result = await db.execute(
        select(TransactionReport)
        .options(selectinload(TransactionReport.customer_report))
        .order_by(TransactionReport.date.desc())
    )
    reports = result.scalars().all()

    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "template", "TransactionReport.xlsx"
    )
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template file not found")

    output = await asyncio.to_thread(_build_transaction_report_excel, reports, template_path)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=TransactionReport.xlsx"}
    )


def _build_transaction_report_excel(reports, template_path) -> BytesIO:
    wb = load_workbook(template_path)
    ws = wb["Data Transaksi "]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for idx, report in enumerate(reports, start=1):
        row = idx + 3  # data starts at row 4 (row 3 is header)
        cr = report.customer_report  # linked CustomerReport via FK

        reg_number = cr.registration_number if cr else None
        sender_name = report.sender_name or (cr.name if cr else None)

        values = [
            report.date,                                                        # A: Tanggal Transaksi
            sender_name,                                                        # B: Nama Pengirim
            reg_number,                                                         # C: No. Id Pengirim
            report.fund_source,                                                 # D: Sumber Dana
            report.transaction_purpose,                                         # E: Tujuan Transaksi
            report.transaction_method,                                          # F: Metode Transaksi
            report.transaction_type,                                            # G: Jenis Transaksi
            report.currency,                                                    # H: Mata Uang
            report.amount,                                                      # I: Nominal Pengiriman
            report.sender_address or (cr.address if cr else None),              # J: Alamat Pengirim
            f"{reg_number} ({sender_name})" if reg_number and sender_name else None,  # K: Id Pengirim (concat)
            report.recipient_name,                                              # L: Nama Penerima
            report.recipient_address,                                           # M: Alamat Penerима
            report.transfer_fee,                                                # N: Biaya Pengiriman
            cr.code_type if cr else None,                                       # O: Kode MC/Bank
            cr.occupation if cr else None,                                      # P: Pekerjaan
            report.risk_level,                                                  # Q: Tingkat Resiko
            "Ya" if report.dttot_check else "Tidak",                            # R: DTTOT
            "Ya" if report.dpppspm_check else "Tidak",                          # S: DPPPSPM
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            # Number format for amount and transfer_fee columns
            if col_idx == 9 or col_idx == 14:
                cell.number_format = '#,##0'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/transaction-report", response_model=TransactionReportDto, status_code=status.HTTP_201_CREATED)
async def create_transaction_report(
    data: TransactionReportCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Создать новый отчет по транзакции"""
    existing = await db.execute(
        select(TransactionReport).where(TransactionReport.transaction_id == data.transaction_id)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Transaction report with transaction_id '{data.transaction_id}' already exists"
        )
    
    report = TransactionReport(
        **data.model_dump(),
        created_date=datetime.utcnow(),
        created_by=current_user.username
    )
    db.add(report)
    
    audit_log = AuditLog(
        entity="transaction_reports",
        entity_id=None,
        action="CREATE",
        old_value=None,
        new_value=json.dumps(data.model_dump(), ensure_ascii=False, default=str),
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.commit()
    await db.refresh(report)
    
    audit_log.entity_id = str(report.id)
    await db.commit()
    
    return TransactionReportDto.model_validate(report, from_attributes=True)


@router.put("/transaction-report/{report_id}", response_model=TransactionReportDto)
async def update_transaction_report(
    report_id: int,
    data: TransactionReportUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Обновить отчет по транзакции"""
    result = await db.execute(
        select(TransactionReport).where(TransactionReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Transaction report with id {report_id} not found"
        )
    
    if data.transaction_id and data.transaction_id != report.transaction_id:
        existing = await db.execute(
            select(TransactionReport).where(
                TransactionReport.transaction_id == data.transaction_id,
                TransactionReport.id != report_id
            )
        )
        if existing.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Transaction report with transaction_id '{data.transaction_id}' already exists"
            )
    
    old_value = {
        "transaction_id": report.transaction_id,
        "date": report.date.isoformat() if report.date else None,
        "sender_name": report.sender_name,
        "recipient_name": report.recipient_name,
        "amount": str(report.amount) if report.amount else None,
        "currency": report.currency,
        "transaction_type": report.transaction_type,
        "risk_level": report.risk_level
    }
    
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if field == "date" and isinstance(value, str):
            from datetime import date as date_type
            value = date_type.fromisoformat(value)
        setattr(report, field, value)
    
    report.updated_at = datetime.utcnow()
    report.updated_by = current_user.username
    
    new_value = {
        "transaction_id": report.transaction_id,
        "date": report.date.isoformat() if report.date else None,
        "sender_name": report.sender_name,
        "recipient_name": report.recipient_name,
        "amount": str(report.amount) if report.amount else None,
        "currency": report.currency,
        "transaction_type": report.transaction_type,
        "risk_level": report.risk_level
    }
    
    audit_log = AuditLog(
        entity="transaction_reports",
        entity_id=str(report_id),
        action="UPDATE",
        old_value=json.dumps(old_value, ensure_ascii=False, default=str),
        new_value=json.dumps(new_value, ensure_ascii=False, default=str),
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.commit()
    await db.refresh(report)
    
    return TransactionReportDto.model_validate(report, from_attributes=True)


@router.delete("/transaction-report/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction_report(
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Удалить отчет по транзакции"""
    result = await db.execute(
        select(TransactionReport).where(TransactionReport.id == report_id)
    )
    report = result.scalar_one_or_none()
    
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Transaction report with id {report_id} not found"
        )
    
    old_value = {
        "id": report.id,
        "transaction_id": report.transaction_id,
        "date": report.date.isoformat() if report.date else None,
        "sender_name": report.sender_name,
        "amount": str(report.amount) if report.amount else None
    }
    
    audit_log = AuditLog(
        entity="transaction_reports",
        entity_id=str(report_id),
        action="DELETE",
        old_value=json.dumps(old_value, ensure_ascii=False, default=str),
        new_value=None,
        created_by=current_user.username,
        created_at=datetime.utcnow()
    )
    db.add(audit_log)
    
    await db.delete(report)
    await db.commit()
    
    return None
