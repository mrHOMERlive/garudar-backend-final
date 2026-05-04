"""
Импорт записей DTTOT/DPPSPM/DPRK/IRAN из Excel файла (форматы PPATK).
Поддерживает два формата:
- Лист "Export" (8 колонок) — DTTOT файл
- Листы "Individu" + "Entitas" — DPRK и Iran файлы
"""
import logging
import re
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.models import Entry

logger = logging.getLogger("garudar_api")


def _str(val, max_len: int = 0) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    if max_len > 0 and len(s) > max_len:
        s = s[:max_len]
    return s


def _derive_source_list(code: str) -> str:
    """Выводит source_list из префикса Kode Densus."""
    c = (code or "").upper().strip()
    if c.startswith("IDD") or c.startswith("EDD"):
        return "DTTOT"
    if c.startswith("ILQ") or c.startswith("ELQ"):
        return "UN-AQ"
    if c.startswith("ILD") or c.startswith("ELD"):
        return "DPPSPM"
    return "DTTOT"


_RE_ALIAS = re.compile(r"\s+alias\s+", re.IGNORECASE)
_RE_NIK = re.compile(r"NIK\s*(?:nomor:?\s*)?(\d{10,16})", re.IGNORECASE)
_RE_PASSPORT = re.compile(r"[Pp]aspor\s*(?:nomor:?\s*)?([A-Z0-9]{5,15})", re.IGNORECASE)


def _parse_export_sheet(ws, source_override: str | None = None) -> list[Entry]:
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        nama = _str(row[0], 0)
        if not nama:
            continue

        deskripsi = _str(row[1], 0)
        terduga = _str(row[2], 255) or ""
        kode = _str(row[3], 255)
        tempat_lahir = _str(row[4], 255)
        tanggal_lahir = _str(row[5], 255)
        wn = _str(row[6], 255)
        alamat = _str(row[7], 4096)

        # --- entry_type ---
        entry_type = "Individual" if "orang" in terduga.lower() else "Entity"

        # --- source_list ---
        if source_override:
            source_list = source_override
        elif kode:
            source_list = _derive_source_list(kode)
        else:
            source_list = "DTTOT"

        # --- names & aliases (split by " alias ") ---
        parts = _RE_ALIAS.split(nama)
        full_name = _str(parts[0], 255)
        aliases = [_str(p, 255) for p in parts[1:] if _str(p)]

        alias_str = "; ".join(a for a in aliases if a) or None
        name1 = full_name
        name2 = aliases[0] if len(aliases) > 0 else None
        name3 = aliases[1] if len(aliases) > 1 else None
        name4 = aliases[2] if len(aliases) > 2 else None

        # --- extract NIK / passport from Deskripsi ---
        identity_no = None
        passport_no = None
        if deskripsi:
            m = _RE_NIK.search(deskripsi)
            if m:
                identity_no = _str(m.group(1), 255)
            m = _RE_PASSPORT.search(deskripsi)
            if m:
                passport_no = _str(m.group(1), 255)

        entries.append(Entry(
            source_list=source_list,
            entry_type=entry_type,
            full_name=_str(full_name, 255),
            name1=_str(name1, 255),
            name2=name2,
            name3=name3,
            name4=name4,
            tittle=_str(kode, 255),
            alias=_str(alias_str, 1024),
            dob=tanggal_lahir,
            pob=tempat_lahir,
            nationality=wn,
            passport_no=passport_no,
            identity_no=identity_no,
            address=alamat,
            additional_info=_str(deskripsi, 4096),
            load_date=date.today(),
        ))
    return entries


def _na(val) -> str | None:
    """Преобразует 'NA' / None в None."""
    s = _str(val, 0)
    if s and s.upper() == "NA":
        return None
    return s


def _parse_individu_entitas_sheets(wb, source_override: str) -> list[Entry]:
    """Парсит DPRK/Iran формат: листы 'Individu' и 'Entitas'."""
    entries = []

    # --- Individu (Individual) ---
    ws_ind = wb["Individu"] if "Individu" in wb.sheetnames else None
    if ws_ind:
        for row in ws_ind.iter_rows(min_row=3, values_only=True):  # row 1=header, row 2=section label
            if not row or not row[1]:
                continue
            full_name = _str(row[1], 255)
            if not full_name:
                continue

            # Alias 1..N — между col 6 и Kewarganegaraan (nationality)
            # Find first non-alias column by looking for known header keywords
            # Col layout: Ref, Nama, Gelar, Pekerjaan, TglLahir, TmptLahir, Alias1..N, Kewarganeg, Paspor, Identitas, Alamat, Info
            # Count how many alias columns by finding the nationality column
            # Simplest: try to find the nationality column (usually has country names)
            # For DPRK: 5 aliases (cols 6-10), then nationality=11, paspor=12, id=13, alamat=14, info=15
            # For Iran: more aliases, so we detect dynamically
            aliases = []
            nationality = None
            passport_no = None
            identity_no = None
            address = None
            additional_info = None
            # Collect all non-None values from col 6 onwards into a list
            remaining = [_na(row[i]) for i in range(6, len(row)) if i < len(row)]
            # The last 4 columns (before info) are: nationality, passport, identity, address
            # and last is info. Aliases are everything before those.
            # Detect: find index where we transition from alias-like values to structured fields
            # Strategy: collect aliases until we hit a value with country-like length or structured format
            # Simpler: use fixed offsets from the END since the structure is consistent
            if len(remaining) >= 5:
                additional_info = _str(remaining[-1], 4096)
                address = _str(remaining[-2], 4096)
                identity_no = _str(remaining[-3], 255)
                passport_no = _str(remaining[-4], 255)
                nationality = _str(remaining[-5], 255)
                aliases = [v for v in remaining[:-5] if v]
            elif len(remaining) >= 1:
                additional_info = _str(remaining[-1], 4096)
                aliases = [v for v in remaining[:-1] if v]

            alias_str = "; ".join(aliases) if aliases else None
            entries.append(Entry(
                source_list=source_override,
                entry_type="Individual",
                full_name=full_name,
                name1=full_name,
                name2=_str(aliases[0], 255) if len(aliases) > 0 else None,
                name3=_str(aliases[1], 255) if len(aliases) > 1 else None,
                name4=_str(aliases[2], 255) if len(aliases) > 2 else None,
                tittle=_str(row[0], 255),
                job_title=_str(_na(row[3]), 255) if len(row) > 3 else None,
                dob=_na(row[4]) if len(row) > 4 else None,
                pob=_na(row[5]) if len(row) > 5 else None,
                alias=_str(alias_str, 1024),
                nationality=nationality,
                passport_no=passport_no,
                identity_no=identity_no,
                address=address,
                additional_info=additional_info,
                load_date=date.today(),
            ))

    # --- Entitas (Entity) ---
    ws_ent = wb["Entitas"] if "Entitas" in wb.sheetnames else None
    if ws_ent:
        for row in ws_ent.iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue
            full_name = _str(row[1], 255)
            if not full_name:
                continue

            # Aliases from col 2 until address columns, addresses from later cols, info at end
            remaining = [_na(row[i]) for i in range(2, len(row)) if i < len(row)]
            additional_info = None
            addresses = []
            aliases = []
            if remaining:
                additional_info = _str(remaining[-1], 4096)
                # Look for address columns (typically last 3-4 before info)
                addr_count = 0
                for v in reversed(remaining[:-1]):
                    if v and addr_count < 4:
                        addresses.insert(0, v)
                        addr_count += 1
                    else:
                        break
                aliases = [v for v in remaining[:len(remaining)-1-addr_count] if v]

            alias_str = "; ".join(aliases) if aliases else None
            address = "; ".join(addresses) if addresses else None

            entries.append(Entry(
                source_list=source_override,
                entry_type="Entity",
                full_name=full_name,
                name1=full_name,
                name2=_str(aliases[0], 255) if len(aliases) > 0 else None,
                name3=_str(aliases[1], 255) if len(aliases) > 1 else None,
                name4=_str(aliases[2], 255) if len(aliases) > 2 else None,
                tittle=_str(row[0], 255),
                alias=_str(alias_str, 1024),
                address=_str(address, 4096),
                additional_info=additional_info,
                load_date=date.today(),
            ))

    return entries


def parse_excel(file_bytes: bytes, source_override: str | None = None) -> list[Entry]:
    """Парсит PPATK Excel файл, автоматически определяя формат."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    if "Export" in wb.sheetnames:
        entries = _parse_export_sheet(wb["Export"], source_override)
        logger.info(f"Parsed {len(entries)} entries from Export sheet (source_override={source_override})")
        wb.close()
        return entries

    if "Individu" in wb.sheetnames or "Entitas" in wb.sheetnames:
        if not source_override:
            logger.warning("Individu/Entitas format detected but no source_override set")
        entries = _parse_individu_entitas_sheets(wb, source_override or "UNKNOWN")
        logger.info(f"Parsed {len(entries)} entries from Individu/Entitas sheets (source={source_override})")
        wb.close()
        return entries

    logger.warning(f"Unknown Excel format. Available sheets: {wb.sheetnames}")
    wb.close()
    return []


async def import_entries_from_bytes(file_bytes: bytes, source_lists_to_delete: list[str],
                                     source_override: str | None, db) -> int:
    """Парсит Excel и заменяет записи для указанных source_list значений."""
    from sqlalchemy import delete
    entries = parse_excel(file_bytes, source_override)
    if not entries:
        logger.warning("No entries parsed from Excel file")
        return 0

    await db.execute(delete(Entry).where(Entry.source_list.in_(source_lists_to_delete)))
    db.add_all(entries)
    await db.commit()

    logger.info(f"Imported {len(entries)} entries, replaced source_lists={source_lists_to_delete}")
    return len(entries)
